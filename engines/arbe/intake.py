# -*- coding: utf-8 -*-
"""Deterministic CR60 analysis intake and source-binding discovery.

The intake layer is deliberately separate from Pi and from any build/debug
side effects.  It reads user-provided materials, records provenance for every
candidate value, and refuses to guess a software/vehicle/COEM binding when
the materials disagree or do not contain enough information.

This is the first control-plane artifact consumed by later modules:
``cr60-analysis-intake.v1``.  It is safe to run before the target Linux
workspace is touched.  It does not checkout, copy, compile, start ROS, or
attach GDB.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
from pathlib import Path
from typing import Any, Iterable


SCHEMA_VERSION = "cr60-analysis-intake.v1"
SUPPORTED_MATERIAL_SUFFIXES = {
    ".csv",
    ".json",
    ".md",
    ".txt",
    ".xlsx",
    ".yaml",
    ".yml",
}

# These aliases are intentionally about *field names*, not feature names.
# Values are still taken from the supplied material and never from path-name
# guesses.  The spreadsheet contract reflects the current CR60 problem-list
# workbook: B/Ticket, C/trigger function, E/vehicle, G/trigger version,
# J/data path.
FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "ticket_id": (
        "ticket",
        "ticketid",
        "ticketno",
        "tr",
        "trno",
        "问题单",
        "问题单号",
        "单号",
    ),
    "function": (
        "function",
        "feature",
        "warningfunction",
        "triggerfunction",
        "触发功能",
        "功能",
        "报警功能",
    ),
    "vehicle": (
        "vehicle",
        "vehicleproject",
        "car",
        "model",
        "车型",
        "车型项目",
    ),
    "customer": ("customer", "oem", "客户", "主机厂"),
    "coem": (
        "coem",
        "coemproject",
        "coemprojectdir",
        "coem工程",
        "coem项目",
    ),
    "software_version": (
        "softwareversion",
        "triggerversion",
        "version",
        "codeversion",
        "softwarebranch",
        "触发版本",
        "软件版本",
        "代码版本",
        "版本",
    ),
    "code_branch": (
        "branch",
        "codebranch",
        "algobranch",
        "sourcebranch",
        "代码分支",
        "子仓分支",
        "分支",
    ),
    "data_path": (
        "datapath",
        "datafile",
        "bagpath",
        "recordpath",
        "录制数据",
        "数据路径",
        "数据文件",
        "bag",
    ),
    "dbc": ("dbc", "dbcpath", "can数据库", "can数据库路径"),
    "cuda_sheet": (
        "cudasheet",
        "xlsxsheet",
        "cudaconfigsheet",
        "车型sheet",
        "车型表",
    ),
    "code_root": ("coderoot", "source_root", "源码仓", "代码仓"),
    "algo_source_root": ("algosourceroot", "algo_source", "算法子仓"),
    "arbe_root": ("arberoot", "arbe_root", "仿真仓", "arbe仓"),
}

SINGULAR_FIELDS = {
    "ticket_id",
    "vehicle",
    "customer",
    "coem",
    "software_version",
    "code_branch",
    "dbc",
    "code_root",
    "arbe_root",
}
REQUIRED_BINDING_FIELDS = ("software_version", "vehicle", "coem")


def _norm_key(value: object) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", str(value or "").lower())


_NORMALIZED_ALIASES = {
    field: {_norm_key(alias) for alias in aliases}
    for field, aliases in FIELD_ALIASES.items()
}


def _field_for_key(key: object) -> str | None:
    normalized = _norm_key(key)
    if not normalized:
        return None
    for field, aliases in _NORMALIZED_ALIASES.items():
        if normalized in aliases:
            return field
    return None


def _clean_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip().strip("\"'")


def _value_key(value: object) -> str:
    return re.sub(r"\s+", " ", _clean_value(value)).strip().casefold()


def _split_values(value: object) -> list[str]:
    if isinstance(value, (list, tuple, set)):
        result: list[str] = []
        for item in value:
            result.extend(_split_values(item))
        return result
    cleaned = _clean_value(value)
    if not cleaned:
        return []
    # A path may contain commas on some systems, so only split delimiters when
    # the value clearly represents a short list rather than a filesystem path.
    if "\n" in cleaned:
        return [part.strip() for part in cleaned.splitlines() if part.strip()]
    if re.search(r"[,;]", cleaned) and not re.search(r"[/\\]", cleaned):
        return [part.strip() for part in re.split(r"[,;]", cleaned) if part.strip()]
    return [cleaned]


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _selected_value(fields: dict[str, Any], field: str, default: str = "") -> str:
    value = fields.get(field, {}).get("value")
    if isinstance(value, list):
        return str(value[0]).strip() if value else default
    return str(value or default).strip()


def _path_basename(path_text: str) -> str:
    return str(path_text).replace("\\", "/").rstrip("/").rsplit("/", 1)[-1]


def _path_parent(path_text: str) -> str:
    normalized = str(path_text).replace("\\", "/").rstrip("/")
    return normalized.rsplit("/", 1)[0] if "/" in normalized else ""


def _stable_handoff_id(
    *,
    data_paths: list[str],
    material_records: list[dict[str, Any]],
    identity: dict[str, Any],
    source_context: dict[str, Any],
) -> str:
    material_fingerprint = [
        {"path": item.get("path", ""), "sha256": item.get("sha256", "")}
        for item in material_records
    ]
    canonical = json.dumps(
        {
            "data_paths": data_paths,
            "materials": material_fingerprint,
            "identity": identity,
            "source_context": source_context,
        },
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    return "intake-" + hashlib.sha256(canonical.encode("utf-8")).hexdigest()[:16]


def _candidate(
    field: str,
    value: object,
    *,
    source: str,
    locator: str,
    method: str,
    priority: int,
    authoritative: bool,
) -> dict[str, Any] | None:
    cleaned = _clean_value(value)
    if not cleaned:
        return None
    return {
        "field": field,
        "value": cleaned,
        "source": source,
        "locator": locator,
        "method": method,
        "priority": priority,
        "authoritative": authoritative,
    }


def _append_candidate(
    candidates: list[dict[str, Any]], candidate: dict[str, Any] | None
) -> None:
    if candidate is None:
        return
    identity = (
        candidate["field"],
        _value_key(candidate["value"]),
        candidate["source"],
        candidate["locator"],
    )
    if any(
        (
            item["field"],
            _value_key(item["value"]),
            item["source"],
            item["locator"],
        )
        == identity
        for item in candidates
    ):
        return
    candidates.append(candidate)


def _iter_mapping_values(
    value: object, *, path: tuple[str, ...] = ()
) -> Iterable[tuple[str, str, object]]:
    if isinstance(value, list):
        for index, item in enumerate(value):
            if isinstance(item, (dict, list)):
                yield from _iter_mapping_values(item, path=path + (str(index),))
        return
    if isinstance(value, dict):
        for key, child in value.items():
            key_text = str(key)
            child_path = path + (key_text,)
            field = _field_for_key(key_text)
            if field is not None and not isinstance(child, (dict, list, tuple)):
                yield field, ".".join(child_path), child
            elif field is not None and isinstance(child, (list, tuple)):
                for item in child:
                    if not isinstance(item, (dict, list, tuple)):
                        yield field, ".".join(child_path), item
            if isinstance(child, dict):
                yield from _iter_mapping_values(child, path=child_path)
            elif isinstance(child, list):
                for index, item in enumerate(child):
                    if isinstance(item, dict):
                        yield from _iter_mapping_values(
                            item, path=child_path + (str(index),)
                        )


def _parse_structured_material(
    path: Path, *, source: str, candidates: list[dict[str, Any]], errors: list[str]
) -> bool:
    try:
        text = path.read_text(encoding="utf-8-sig")
    except (OSError, UnicodeError) as exc:
        errors.append(f"{path}: read failed: {type(exc).__name__}: {exc}")
        return False

    data: object
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        data = None
    if data is None and path.suffix.lower() in {".yaml", ".yml"}:
        try:
            import yaml  # type: ignore

            data = yaml.safe_load(text)
        except ImportError:
            data = None
        except Exception as exc:  # noqa: BLE001 - material is external input
            errors.append(f"{path}: YAML parse failed: {type(exc).__name__}: {exc}")
            return False

    if isinstance(data, (dict, list)):
        for field, locator, value in _iter_mapping_values(data):
            for item in _split_values(value):
                _append_candidate(
                    candidates,
                    _candidate(
                        field,
                        item,
                        source=source,
                        locator=locator,
                        method="structured_key",
                        priority=80,
                        authoritative=True,
                    ),
                )
        return True

    # JSON may be a scalar or malformed YAML may be unavailable.  A small
    # line parser still handles hand-written handoff notes without treating
    # arbitrary prose as a binding.
    parsed_any = False
    for line_number, line in enumerate(text.splitlines(), start=1):
        match = re.match(
            r"^\s*(?:[-*]\s*)?(?:\|\s*)?([^:=|]+?)\s*[:=]\s*([^|#]+?)\s*(?:\|\s*)?$",
            line,
        )
        if not match:
            continue
        field = _field_for_key(match.group(1))
        if field is None:
            continue
        parsed_any = True
        for item in _split_values(match.group(2)):
            _append_candidate(
                candidates,
                _candidate(
                    field,
                    item,
                    source=source,
                    locator=f"line:{line_number}",
                    method="key_value_line",
                    priority=60,
                    authoritative=False,
                ),
            )
    return parsed_any


def _excel_column_name(index: int) -> str:
    name = ""
    current = index + 1
    while current:
        current, remainder = divmod(current - 1, 26)
        name = chr(65 + remainder) + name
    return name


def _xlsx_match_tokens(
    data_paths: list[str], ticket_id: str = "", match_text: list[str] | None = None
) -> list[str]:
    raw = list(match_text or []) + ([ticket_id] if ticket_id else [])
    for path_text in data_paths:
        raw.append(path_text)
        raw.append(Path(path_text).name)
        raw.append(Path(path_text).stem)
    tokens: list[str] = []
    for item in raw:
        normalized = _value_key(item)
        if len(normalized) < 3:
            continue
        if normalized not in tokens:
            tokens.append(normalized)
    return tokens


def _row_looks_like_header(row: tuple[Any, ...], mapping: dict[int, str]) -> bool:
    """Return whether a fallback-contract first row is a header row."""
    hits = 0
    for index in mapping:
        if index >= len(row):
            continue
        cell = _clean_value(row[index])
        if not cell:
            continue
        if _field_for_key(cell) is not None:
            hits += 1
            continue
        if _norm_key(cell) in {
            "ticketno",
            "triggerfunction",
            "vehicle",
            "triggerversion",
            "datapath",
        }:
            hits += 1
    return hits >= 2


def _parse_xlsx_material(
    path: Path,
    *,
    source: str,
    candidates: list[dict[str, Any]],
    errors: list[str],
    data_paths: list[str],
    ticket_id: str,
    match_text: list[str] | None,
) -> dict[str, Any]:
    metadata: dict[str, Any] = {"path": str(path), "sheets": [], "matched_rows": 0}
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        errors.append(f"{path}: openpyxl is not installed")
        metadata["status"] = "unreadable_missing_dependency"
        return metadata

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - workbook is external input
        errors.append(f"{path}: XLSX parse failed: {type(exc).__name__}: {exc}")
        metadata["status"] = "unreadable"
        return metadata

    tokens = _xlsx_match_tokens(data_paths, ticket_id, match_text)
    # Actual CR60 problem-list fallback when a workbook has no usable header.
    fallback_columns = {
        1: "ticket_id",  # B
        2: "function",  # C
        4: "vehicle",  # E
        6: "software_version",  # G
        9: "data_path",  # J
    }

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(min_row=1, max_row=2000, values_only=True))
        nonempty_rows = [row for row in rows if any(_clean_value(cell) for cell in row)]
        if not nonempty_rows:
            continue
        header_index: int | None = None
        header_map: dict[int, str] = {}
        for row_index, row in enumerate(nonempty_rows[:30]):
            detected = {
                index: field
                for index, cell in enumerate(row)
                if (field := _field_for_key(cell)) is not None
            }
            if len(detected) >= 2:
                header_index = row_index
                header_map = detected
                break
        mapping = header_map or fallback_columns
        data_start = (header_index + 1) if header_index is not None else 0
        if header_index is None and _row_looks_like_header(nonempty_rows[0], mapping):
            data_start = 1
        data_rows = nonempty_rows[data_start:]
        matched: list[tuple[int, tuple[Any, ...]]] = []
        for row_index, row in enumerate(data_rows, start=data_start + 1):
            row_text = " ".join(_value_key(cell) for cell in row if _clean_value(cell))
            if not tokens or any(token in row_text for token in tokens):
                matched.append((row_index, row))
        # Do not ingest an entire issue list when the requested record could
        # not be identified.  A one-row workbook remains useful without a
        # match token; a multi-row workbook is explicitly reported unmatched.
        if not tokens and len(data_rows) != 1:
            matched = []
        if tokens and not matched:
            metadata["sheets"].append(
                {"name": worksheet.title, "status": "no_matching_row"}
            )
            continue

        sheet_info = {
            "name": worksheet.title,
            "status": "matched",
            "header_row": (header_index + 1) if header_index is not None else None,
            "mapping": {
                _excel_column_name(index): field for index, field in mapping.items()
            },
            "rows": [row_index for row_index, _ in matched],
        }
        metadata["sheets"].append(sheet_info)
        for row_index, row in matched:
            metadata["matched_rows"] += 1
            for column_index, field in mapping.items():
                if column_index >= len(row):
                    continue
                for item in _split_values(row[column_index]):
                    _append_candidate(
                        candidates,
                        _candidate(
                            field,
                            item,
                            source=source,
                            locator=(
                                f"{worksheet.title}!"
                                f"{_excel_column_name(column_index)}{row_index}"
                            ),
                            method=(
                                "xlsx_header" if header_map else "xlsx_column_contract"
                            ),
                            priority=75 if header_map else 70,
                            authoritative=True,
                        ),
                    )
    workbook.close()
    metadata["status"] = "matched" if metadata["matched_rows"] else "no_match"
    return metadata


def _parse_csv_material(
    path: Path, *, source: str, candidates: list[dict[str, Any]], errors: list[str]
) -> bool:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            parsed = False
            for row_number, row in enumerate(reader, start=2):
                for key, value in row.items():
                    field = _field_for_key(key)
                    if field is None:
                        continue
                    parsed = True
                    for item in _split_values(value):
                        _append_candidate(
                            candidates,
                            _candidate(
                                field,
                                item,
                                source=source,
                                locator=f"line:{row_number}/{key}",
                                method="csv_header",
                                priority=70,
                                authoritative=True,
                            ),
                        )
            return parsed
    except (OSError, UnicodeError, csv.Error) as exc:
        errors.append(f"{path}: CSV parse failed: {type(exc).__name__}: {exc}")
        return False


def _collect_material_files(material_paths: list[str]) -> tuple[list[Path], list[str]]:
    files: list[Path] = []
    missing: list[str] = []
    seen: set[str] = set()
    for raw in material_paths:
        value = str(raw or "").strip()
        if not value:
            continue
        path = Path(value).expanduser()
        if not path.exists():
            missing.append(value)
            continue
        if path.is_file():
            if path.suffix.lower() in SUPPORTED_MATERIAL_SUFFIXES:
                resolved = str(path.resolve())
                if resolved not in seen:
                    files.append(path.resolve())
                    seen.add(resolved)
            continue
        if path.is_dir():
            for child in sorted(path.rglob("*"), key=lambda item: str(item).lower()):
                if not child.is_file() or child.suffix.lower() not in SUPPORTED_MATERIAL_SUFFIXES:
                    continue
                resolved = str(child.resolve())
                if resolved not in seen:
                    files.append(child.resolve())
                    seen.add(resolved)
    return files, missing


def _path_record(path_text: str) -> dict[str, Any]:
    path = Path(path_text).expanduser()
    record: dict[str, Any] = {
        "path": path_text,
        "local_validation": "not_found",
        "kind": "unknown",
        "data_file_count": None,
    }
    try:
        if path.is_file():
            record.update({"local_validation": "exists", "kind": "file", "data_file_count": 1})
        elif path.is_dir():
            record.update(
                {
                    "local_validation": "exists",
                    "kind": "directory",
                    "data_file_count": sum(
                        1
                        for child in path.rglob("*")
                        if child.is_file() and child.suffix.lower() in {".bag", ".blf", ".mf4", ".db"}
                    ),
                }
            )
    except OSError as exc:
        record["local_validation"] = f"error:{type(exc).__name__}"
    # Linux paths supplied from a Windows control plane are expected to be
    # unverified locally.  A later SSH data-prep module performs the remote
    # existence and checksum check.
    if path_text.startswith("/") or path_text.startswith("\\"):
        record["remote_candidate"] = True
        if record["local_validation"] == "not_found":
            record["local_validation"] = "remote_unverified"
    else:
        record["remote_candidate"] = False
    return record


def _select_fields(candidates: list[dict[str, Any]]) -> tuple[dict[str, Any], list[dict[str, Any]], list[str]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for candidate in candidates:
        grouped.setdefault(candidate["field"], []).append(candidate)

    fields: dict[str, Any] = {}
    conflicts: list[dict[str, Any]] = []
    warnings: list[str] = []
    all_fields = set(FIELD_ALIASES) | {"server_host", "server_user"}
    for field in sorted(all_fields):
        items = sorted(
            grouped.get(field, []),
            key=lambda item: (-int(item.get("priority", 0)), item["source"], item["locator"]),
        )
        unique: dict[str, dict[str, Any]] = {}
        for item in items:
            unique.setdefault(_value_key(item["value"]), item)
        values = list(unique.values())
        if not values:
            fields[field] = {
                "status": "missing",
                "value": [] if field in {"function", "data_path"} else None,
                "candidates": [],
            }
            continue
        if field in {"function", "data_path"}:
            selected = [item["value"] for item in values]
            fields[field] = {
                "status": "resolved",
                "value": selected,
                "selected_from": values[0],
                "candidates": items,
            }
            continue
        path_hint_only = all(item.get("method") == "path_hint" for item in values)
        if path_hint_only:
            fields[field] = {
                "status": "hint_only",
                "value": None,
                "candidates": items,
            }
            warnings.append(f"{field}_path_hint_not_authoritative")
            continue
        if len(values) > 1:
            conflict = {"field": field, "candidates": values}
            conflicts.append(conflict)
            fields[field] = {
                "status": "conflict",
                "value": None,
                "candidates": items,
            }
            continue
        fields[field] = {
            "status": "resolved",
            "value": values[0]["value"],
            "selected_from": values[0],
            "candidates": items,
        }
    return fields, conflicts, warnings


def build_intake(
    *,
    data_paths: list[str] | None = None,
    material_paths: list[str] | None = None,
    match_text: list[str] | None = None,
    software_version: str = "",
    vehicle: str = "",
    customer: str = "",
    coem: str = "",
    code_branch: str = "",
    ticket_id: str = "",
    function: list[str] | None = None,
    server_host: str = "",
    server_user: str = "",
    server_port: int = 22,
    arbe_root: str = "",
    algo_source_root: str = "",
    code_root: str = "",
    dbc: str = "",
    cuda_sheet: str = "",
    customer_claim: str = "",
    preferred_radar: str = "auto",
) -> dict[str, Any]:
    """Build a provenance-preserving intake payload without side effects."""

    explicit: dict[str, str] = {
        "software_version": software_version,
        "vehicle": vehicle,
        "customer": customer,
        "coem": coem,
        "code_branch": code_branch,
        "ticket_id": ticket_id,
        "server_host": server_host,
        "server_user": server_user,
        "arbe_root": arbe_root,
        "code_root": code_root,
        "dbc": dbc,
        "algo_source_root": algo_source_root,
        "cuda_sheet": cuda_sheet,
    }
    data_values = [str(item).strip() for item in (data_paths or []) if str(item).strip()]
    material_values = [str(item).strip() for item in (material_paths or []) if str(item).strip()]
    candidates: list[dict[str, Any]] = []
    for field, value in explicit.items():
        if not value:
            continue
        _append_candidate(
            candidates,
            _candidate(
                field,
                value,
                source="explicit_input",
                locator=f"input.{field}",
                method="explicit",
                priority=100,
                authoritative=True,
            ),
        )
    for item in function or []:
        for value in _split_values(item):
            _append_candidate(
                candidates,
                _candidate(
                    "function",
                    value,
                    source="explicit_input",
                    locator="input.function",
                    method="explicit",
                    priority=100,
                    authoritative=True,
                ),
            )
    for path_text in data_values:
        _append_candidate(
            candidates,
            _candidate(
                "data_path",
                path_text,
                source="explicit_input",
                locator="input.data_paths",
                method="explicit",
                priority=100,
                authoritative=True,
            ),
        )

    material_files, missing_materials = _collect_material_files(material_values)
    material_records: list[dict[str, Any]] = []
    errors: list[str] = []
    match_values = list(match_text or [])
    for path in material_files:
        source = str(path)
        record: dict[str, Any] = {
            "path": source,
            "suffix": path.suffix.lower(),
            "sha256": "",
            "status": "not_parsed",
        }
        try:
            record["sha256"] = _sha256(path)
        except OSError as exc:
            errors.append(f"{path}: hash failed: {type(exc).__name__}: {exc}")
        if path.suffix.lower() == ".xlsx":
            record.update(
                _parse_xlsx_material(
                    path,
                    source=source,
                    candidates=candidates,
                    errors=errors,
                    data_paths=data_values,
                    ticket_id=ticket_id,
                    match_text=match_values,
                )
            )
        elif path.suffix.lower() == ".csv":
            record["status"] = (
                "parsed"
                if _parse_csv_material(path, source=source, candidates=candidates, errors=errors)
                else "no_fields"
            )
        else:
            record["status"] = (
                "parsed"
                if _parse_structured_material(
                    path, source=source, candidates=candidates, errors=errors
                )
                else "no_fields"
            )
        material_records.append(record)

    fields, conflicts, selection_warnings = _select_fields(candidates)
    selected_data_paths = [str(item) for item in fields["data_path"].get("value", [])]
    # Explicit data paths are already candidates; material-only paths are
    # accepted only when their row/key was actually matched and authoritative.
    path_records = [_path_record(item) for item in selected_data_paths]
    missing: list[str] = []
    if not selected_data_paths:
        missing.append("data_paths")
    for field in REQUIRED_BINDING_FIELDS:
        if fields[field].get("status") != "resolved" or not fields[field].get("value"):
            missing.append(field)
    if fields["code_branch"].get("status") != "resolved":
        missing.append("code_branch_or_version_to_branch_mapping")

    confirmation_required: list[dict[str, Any]] = []
    for conflict in conflicts:
        confirmation_required.append(
            {
                "type": "conflict",
                "field": conflict["field"],
                "reason": "multiple authoritative materials disagree",
                "candidates": conflict["candidates"],
            }
        )
    for field in missing:
        confirmation_required.append(
            {
                "type": "missing",
                "field": field,
                "reason": "required before safe source/build binding",
            }
        )
    for path_record in path_records:
        if path_record.get("local_validation") == "remote_unverified":
            confirmation_required.append(
                {
                    "type": "remote_validation",
                    "field": "data_paths",
                    "path": path_record["path"],
                    "reason": "path looks remote and needs data-prep checksum validation",
                }
            )

    remote_unverified = any(
        item.get("local_validation") == "remote_unverified" for item in path_records
    )
    if not selected_data_paths:
        intake_status = "blocked_missing_input"
    elif conflicts or missing or remote_unverified:
        intake_status = "needs_confirmation"
    else:
        intake_status = "ready"

    # The detailed intake state is useful to Pi, while the top-level status is
    # intentionally compatible with the upstream/downstream handoff consumer.
    # A handoff is only ``ready`` after all identity/source gates pass; a path
    # that exists only on Linux is ``partial`` until data-prep verifies it.
    handoff_status = (
        "blocked"
        if intake_status == "blocked_missing_input" or conflicts or missing
        else "partial"
        if remote_unverified
        else "ready"
    )
    identity_fields = {
        field: fields[field]
        for field in (
            "ticket_id",
            "function",
            "customer",
            "vehicle",
            "coem",
            "software_version",
            "code_branch",
            "cuda_sheet",
        )
    }
    source_fields = {
        field: fields[field]
        for field in (
            "server_host",
            "server_user",
            "arbe_root",
            "algo_source_root",
            "code_root",
            "dbc",
            "cuda_sheet",
        )
    }
    selected_functions = [str(item) for item in fields["function"].get("value", [])]
    selected_ticket = _selected_value(fields, "ticket_id")
    selected_software = _selected_value(fields, "software_version")
    selected_vehicle = _selected_value(fields, "vehicle")
    selected_customer = _selected_value(fields, "customer")
    selected_coem = _selected_value(fields, "coem")
    selected_branch = _selected_value(fields, "code_branch")
    selected_cuda_sheet = _selected_value(fields, "cuda_sheet")
    selected_server_host = _selected_value(fields, "server_host")
    selected_server_user = _selected_value(fields, "server_user")
    selected_arbe_root = _selected_value(fields, "arbe_root")
    selected_algo_root = _selected_value(fields, "algo_source_root")
    selected_code_root = _selected_value(fields, "code_root")
    selected_dbc = _selected_value(fields, "dbc")

    source_selector: dict[str, Any] = {}
    if selected_branch:
        source_selector["algo_submodule_branch"] = selected_branch
    if selected_software:
        source_selector["software_version"] = selected_software
    if selected_coem:
        source_selector["coem"] = selected_coem
    if selected_vehicle:
        source_selector["vehicle"] = selected_vehicle

    handoff_cases: list[dict[str, Any]] = []
    for index, path_record in enumerate(path_records):
        path_text = str(path_record["path"])
        stem = _path_basename(path_text).rsplit(".", 1)[0]
        case_id = selected_ticket or stem or f"case-{index + 1}"
        if len(path_records) > 1 and selected_ticket:
            case_id = f"{selected_ticket}__{stem or index + 1}"
        suffix = Path(path_text.replace("\\", "/")).suffix.lower().lstrip(".")
        bag_item: dict[str, Any] = {
            "path": path_text,
            "format": suffix,
            "size_bytes": None,
            "sha256": "",
            "local_validation": path_record.get("local_validation", ""),
        }
        try:
            local_path = Path(path_text).expanduser()
            if local_path.is_file():
                bag_item["size_bytes"] = local_path.stat().st_size
                bag_item["sha256"] = _sha256(local_path)
        except OSError:
            pass
        handoff_cases.append(
            {
                "case_id": case_id,
                "tr_id": selected_ticket,
                "data_dir": _path_parent(path_text),
                "bag_paths": [bag_item],
                "functions_hint": selected_functions,
                "customer_claim": customer_claim,
                "preferred_radar": preferred_radar or "auto",
                "source_selector": source_selector,
            }
        )

    environment = {
        "server": {
            "host": selected_server_host,
            "user": selected_server_user,
            "port": int(server_port),
        },
        "arbe": {
            "workspace": selected_arbe_root,
            "algo_source_root": selected_algo_root,
        },
        "vehicle": {
            "customer": selected_customer,
            "model": selected_vehicle,
            "coem": selected_coem,
            "cuda_sheet": selected_cuda_sheet,
        },
        "build": {
            "software_version": selected_software,
            "code_branch": selected_branch,
            "code_root": selected_code_root,
            "dbc": selected_dbc,
        },
    }
    data_root = _path_parent(selected_data_paths[0]) if selected_data_paths else ""
    handoff_id = _stable_handoff_id(
        data_paths=selected_data_paths,
        material_records=material_records,
        identity=identity_fields,
        source_context=source_fields,
    )

    return {
        "schema_version": SCHEMA_VERSION,
        "status": handoff_status,
        "intake_status": intake_status,
        "handoff_id": handoff_id,
        "input_policy": {
            "materials_first": True,
            "path_names_are_not_identity_evidence": True,
            "conflicting_authoritative_values_fail_closed": True,
            "remote_path_validation_deferred_to_data_prep": True,
        },
        "data": {
            "paths": path_records,
            "count": len(path_records),
            "root": data_root,
            "cases": handoff_cases,
        },
        "identity": identity_fields,
        "source_context": source_fields,
        "environment": environment,
        "materials": material_records,
        "material_inputs": {
            "requested": material_values,
            "discovered_files": [str(path) for path in material_files],
            "missing_paths": missing_materials,
        },
        "candidates": candidates,
        "missing": missing,
        "conflicts": conflicts,
        "confirmation_required": confirmation_required,
        "warnings": selection_warnings
        + (["material_paths_missing"] if missing_materials else []),
        "errors": errors,
        "next_actions": [
            "confirm or provide missing identity/source fields",
            "run remote data-prep checksum validation for remote paths",
            "run arbe-preflight against the confirmed workspace",
        ],
    }


__all__ = [
    "SCHEMA_VERSION",
    "SUPPORTED_MATERIAL_SUFFIXES",
    "build_intake",
]
