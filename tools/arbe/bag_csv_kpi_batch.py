#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Standalone KPI batch script for corner radar GT CSV + rosbag.

Features:
- Single pair mode: --bag + --csv
- Folder batch mode: --input-dir (auto match xxx.bag -> xxx_corner_radar_gt.csv, fallback xxx.csv)
- Reads warning topic (default: /corner_radar/warning_status_raw)
- Uses per-radar LGU timestamps (default: /wf/corner_radar/lgu_data_<id>) as event time base
- PM standard metrics (same as FrameSync):
  - TP: pred overlaps GT
  - FP: pred overlaps no GT
  - FN: GT has no pred overlap
  - interruption stats for multi-segment TP
- Exports:
  - *_adas_kpi_summary.csv
  - *_adas_kpi_summary_events.csv
  - *_adas_kpi_report.xlsx
  - batch_kpi_index.csv
"""

import argparse
import csv
import datetime as dt
import math
import os
import sys
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

try:
    import rosbag
except Exception as exc:
    print("[ERROR] Failed to import rosbag. Source ROS env first, e.g. `source /opt/ros/noetic/setup.bash`.", file=sys.stderr)
    raise


K_MAX_RADAR_ID = 4
K_ADAS_WARN_COUNT = 15
K_DEFAULT_EA_THRESHOLD_SEC = 0.25
K_DEFAULT_ED_THRESHOLD_SEC = 0.25
K_DEFAULT_LA_THRESHOLD_SEC = 0.25
K_DEFAULT_LD_THRESHOLD_SEC = 0.25

K_ADAS_LABEL_NAMES = {
    1: "BSD_L",
    2: "BSD_R",
    3: "LCA_L",
    4: "LCA_R",
    5: "DOW_L",
    6: "DOW_R",
    7: "RCW",
    8: "RCTA_L",
    9: "RCTA_R",
    10: "RCTB_L",
    11: "RCTB_R",
    12: "FCTA_L",
    13: "FCTA_R",
    14: "FCTB_L",
    15: "FCTB_R",
}


def radar_direction_from_id(radar_id: int) -> str:
    return {1: "FL", 2: "FR", 3: "RL", 4: "RR"}.get(radar_id, "UNKNOWN")


def trim(s: str) -> str:
    return (s or "").strip()


def normalize_label_key(label: str) -> str:
    out = []
    for c in trim(label):
        o = ord(c)
        if 97 <= o <= 122:
            out.append(chr(o - 32))
        elif (65 <= o <= 90) or (48 <= o <= 57) or c == "_":
            out.append(c)
        elif c in (" ", "-", "/"):
            out.append("_")
    return "".join(out)


def adas_label_index_from_name(label_name: str) -> int:
    key = normalize_label_key(label_name)
    if not key:
        return -1

    alias = {
        "BSD_L": 1,
        "BSD_LEFT": 1,
        "LEFT_BSD": 1,
        "BSD_R": 2,
        "BSD_RIGHT": 2,
        "RIGHT_BSD": 2,
        "LCA_L": 3,
        "LCA_LEFT": 3,
        "LEFT_LCA": 3,
        "LCA_R": 4,
        "LCA_RIGHT": 4,
        "RIGHT_LCA": 4,
        "DOW_L": 5,
        "DOW_LEFT": 5,
        "LEFT_DOW": 5,
        "DOW_R": 6,
        "DOW_RIGHT": 6,
        "RIGHT_DOW": 6,
        "RCW": 7,
        "RCTA_L": 8,
        "RCTA_LEFT": 8,
        "LEFT_RCTA": 8,
        "RCTA_R": 9,
        "RCTA_RIGHT": 9,
        "RIGHT_RCTA": 9,
        "RCTB_L": 10,
        "RCTB_LEFT": 10,
        "LEFT_RCTB": 10,
        "RCTB_R": 11,
        "RCTB_RIGHT": 11,
        "RIGHT_RCTB": 11,
        "FCTA_L": 12,
        "FCTA_LEFT": 12,
        "LEFT_FCTA": 12,
        "FCTA_R": 13,
        "FCTA_RIGHT": 13,
        "RIGHT_FCTA": 13,
        "FCTB_L": 14,
        "FCTB_LEFT": 14,
        "LEFT_FCTB": 14,
        "FCTB_R": 15,
        "FCTB_RIGHT": 15,
        "RIGHT_FCTB": 15,
    }
    return alias.get(key, -1)


@dataclass
class Interval:
    radar_id: int
    label_index: int
    label_name: str
    start_sec: float
    end_sec: float


@dataclass
class CsvMeta:
    csv_path: str
    bag_path: str
    bag_name: str
    bag_start_sec: float
    bag_end_sec: float


@dataclass
class MetricRow:
    radar_id: int
    label_index: int
    label_name: str
    gt_count: int
    pred_count: int
    tp: int
    fn: int
    fp: int
    tpr: float
    fpr: float
    fnr: float
    precision: float
    f1: float
    mean_delay: float
    min_delay: float
    max_delay: float
    mean_overlap: float
    interrupted_tp: int
    interruption_count: int
    interruption_tp_ratio: float
    mean_interrupt_gap: float
    max_interrupt_gap: float
    mean_tp_duration: float
    mean_fn_duration: float
    mean_fp_duration: float
    ea_count: int
    mean_ea_duration: float
    ed_count: int
    mean_ed_duration: float
    la_count: int
    mean_la_duration: float
    ld_count: int
    mean_ld_duration: float
    double_warning_count: int
    mean_double_warning_duration: float


@dataclass
class EventRow:
    radar_id: int
    label_index: int
    label_name: str
    event_type: str
    gt_start: float
    gt_end: float
    pred_start: float
    pred_end: float
    delay: float
    overlap: float
    pred_segments: int
    interruption_count: int
    pred_first_start: float = math.nan
    pred_last_end: float = math.nan
    pred_active_duration: float = math.nan
    pred_span_duration: float = math.nan
    interruption_gaps: List[float] = field(default_factory=list)
    is_interrupted: int = 0
    start_offset: float = math.nan
    end_offset: float = math.nan
    ea_duration: float = math.nan
    ed_duration: float = math.nan
    la_duration: float = math.nan
    ld_duration: float = math.nan
    is_ea: int = 0
    is_ed: int = 0
    is_la: int = 0
    is_ld: int = 0
    double_warning_duration: float = math.nan


@dataclass
class BatchJob:
    bag_path: str
    csv_path: str
    bag_base: str


def parse_float(v: str) -> Optional[float]:
    try:
        return float(trim(v))
    except Exception:
        return None


def parse_int(v: str) -> Optional[int]:
    try:
        return int(trim(v))
    except Exception:
        return None


def fmt6(v: float) -> str:
    return f"{v:.6f}"


def fmt_opt(v: float) -> str:
    return "" if (not math.isfinite(v)) else f"{v:.6f}"


def mean_or_nan(values: List[float]) -> float:
    return (sum(values) / len(values)) if values else math.nan


def str_to_float(v: str) -> float:
    try:
        return float(trim(v))
    except Exception:
        return math.nan


def str_to_int(v: str) -> int:
    try:
        return int(float(trim(v)))
    except Exception:
        return 0


def ros_abs_sec_to_cn_text(v: str) -> str:
    sec = str_to_float(v)
    if not math.isfinite(sec):
        return ""
    try:
        dt_utc = dt.datetime.fromtimestamp(sec, tz=dt.timezone.utc)
        dt_cn = dt_utc.astimezone(dt.timezone(dt.timedelta(hours=8)))
        return dt_cn.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
    except Exception:
        return ""


def is_all_summary_row(r: Dict[str, str]) -> bool:
    return trim(r.get("label_name", "")).upper() == "ALL"


def export_excel_report_from_csv(
    excel_path: str,
    summary_csv_path: str,
    events_csv_path: str,
    report_note: str = "",
) -> None:
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError(f"openpyxl import failed: {exc}")

    def read_csv_rows(path: str) -> Tuple[List[str], List[Dict[str, str]]]:
        with open(path, "r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            rows = [dict(r) for r in reader]
            return list(reader.fieldnames or []), rows

    summary_headers, summary_rows = read_csv_rows(summary_csv_path)
    event_headers, event_rows = read_csv_rows(events_csv_path)
    has_source_bag_cols = ("source_bag_name" in event_headers) or ("source_bag_path" in event_headers)
    source_cols = []
    if "source_bag_name" in event_headers:
        source_cols.append("source_bag_name")
    if "source_bag_path" in event_headers:
        source_cols.append("source_bag_path")
    has_source_bag_cols_summary = ("source_bag_name" in summary_headers) or ("source_bag_path" in summary_headers)
    source_cols_summary = []
    if "source_bag_name" in summary_headers:
        source_cols_summary.append("source_bag_name")
    if "source_bag_path" in summary_headers:
        source_cols_summary.append("source_bag_path")

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary.append(summary_headers)
    for r in summary_rows:
        ws_summary.append([r.get(h, "") for h in summary_headers])

    abs_time_cols = [
        ("gt_start_abs", "gt_start_cn"),
        ("gt_end_abs", "gt_end_cn"),
        ("pred_start_abs", "pred_start_cn"),
        ("pred_end_abs", "pred_end_cn"),
        ("pred_first_start_abs", "pred_first_start_cn"),
        ("pred_last_end_abs", "pred_last_end_cn"),
    ]
    abs_time_map = {src: dst for src, dst in abs_time_cols}

    def interleave_cn_headers(headers: List[str]) -> List[str]:
        out: List[str] = []
        for h in headers:
            out.append(h)
            if h in abs_time_map:
                out.append(abs_time_map[h])
        return out

    def with_cn_time_columns(headers: List[str], row: Dict[str, str]) -> List[str]:
        out: List[str] = []
        for h in headers:
            out.append(row.get(h, ""))
            if h in abs_time_map:
                out.append(ros_abs_sec_to_cn_text(row.get(h, "")))
        return out

    event_headers_ext = interleave_cn_headers(list(event_headers))
    ws_events = wb.create_sheet("Events")
    ws_events.append(event_headers_ext)
    for r in event_rows:
        ws_events.append(with_cn_time_columns(event_headers, r))

    split_headers = source_cols + [
        "radar_id",
        "radar_direction",
        "label_name",
        "event_type",
        "gt_start_abs",
        "gt_end_abs",
        "pred_start_abs",
        "pred_end_abs",
        "pred_first_start_abs",
        "pred_last_end_abs",
        "delay_sec",
        "overlap_sec",
        "gt_duration_sec",
        "pred_duration_sec",
        "pred_active_duration_sec",
        "pred_span_duration_sec",
        "pred_segments",
        "interruption_count",
        "interruption_gaps_sec",
        "is_interrupted",
        "start_offset_sec",
        "end_offset_sec",
        "ea_duration_sec",
        "ed_duration_sec",
        "la_duration_sec",
        "ld_duration_sec",
        "is_ea",
        "is_ed",
        "is_la",
        "is_ld",
        "double_warning_duration_sec",
    ]
    split_headers_ext = interleave_cn_headers(list(split_headers))

    def add_split_sheet(name: str, predicate) -> None:
        ws = wb.create_sheet(name)
        ws.append(split_headers_ext)
        for r in event_rows:
            if predicate(r):
                ws.append(with_cn_time_columns(split_headers, r))

    add_split_sheet("TP", lambda r: trim(r.get("event_type", "")) == "TP")
    add_split_sheet("FP", lambda r: trim(r.get("event_type", "")) == "FP")
    add_split_sheet("FN", lambda r: trim(r.get("event_type", "")) == "FN")
    add_split_sheet("EA", lambda r: str_to_int(r.get("is_ea", "0")) > 0)
    add_split_sheet("ED", lambda r: str_to_int(r.get("is_ed", "0")) > 0)
    add_split_sheet("LA", lambda r: str_to_int(r.get("is_la", "0")) > 0)
    add_split_sheet("LD", lambda r: str_to_int(r.get("is_ld", "0")) > 0)
    add_split_sheet("DoubleWarnings", lambda r: str_to_int(r.get("is_interrupted", "0")) > 0)

    ws_add = wb.create_sheet("AdditionalStats_5G")
    add_headers = [
        "radar_id",
        "radar_direction",
        "label_name",
        "mean_duration_tp_sec",
        "mean_duration_fn_sec",
        "mean_duration_fp_sec",
        "ea_count",
        "mean_duration_ea_sec",
        "ed_count",
        "mean_duration_ed_sec",
        "la_count",
        "mean_duration_la_sec",
        "ld_count",
        "mean_duration_ld_sec",
        "double_warning_count",
        "mean_duration_double_warning_sec",
        "ea_threshold_sec",
        "ed_threshold_sec",
        "la_threshold_sec",
        "ld_threshold_sec",
    ]
    ws_add.append(add_headers)
    for r in summary_rows:
        ws_add.append(
            [
                r.get("radar_id", ""),
                r.get("radar_direction", ""),
                r.get("label_name", ""),
                r.get("mean_tp_duration_sec", ""),
                r.get("mean_fn_duration_sec", ""),
                r.get("mean_fp_duration_sec", ""),
                r.get("ea_count", ""),
                r.get("mean_ea_duration_sec", ""),
                r.get("ed_count", ""),
                r.get("mean_ed_duration_sec", ""),
                r.get("la_count", ""),
                r.get("mean_la_duration_sec", ""),
                r.get("ld_count", ""),
                r.get("mean_ld_duration_sec", ""),
                r.get("double_warning_count", ""),
                r.get("mean_double_warning_duration_sec", ""),
                r.get("ea_threshold_sec", ""),
                r.get("ed_threshold_sec", ""),
                r.get("la_threshold_sec", ""),
                r.get("ld_threshold_sec", ""),
            ]
        )

    ws_metric_dict = wb.create_sheet("KPI_MetricDictionary")
    ws_metric_dict.append(
        [
            "MetricKey(指标键)",
            "ChineseName(中文名称)",
            "Level(层级)",
            "SheetSource(来源子表)",
            "FormulaOrRule(计算公式/判定规则)",
            "Notes(备注)",
        ]
    )
    metric_rows = [
        ("tp", "真阳性数", "Summary", "Summary", "TP = 与GT严格重叠(overlap>0)的GT事件数", "按 radar_id+label_name 聚合"),
        ("fn", "漏报数", "Summary", "Summary", "FN = 无任何预测重叠的GT事件数", "按 radar_id+label_name 聚合"),
        ("fp", "误报数", "Summary", "Summary", "FP = 不与任何GT严格重叠的预测事件数", "按 radar_id+label_name 聚合"),
        ("overlap_sec", "重叠时长", "Event", "Events", "sum(clip(pred,gt)后各段重叠长度)", "仅 TP 有意义"),
        ("pred_segments", "预测分段数", "Event", "Events", "GT内重叠片段合并后的段数", ">=2 可能中断"),
        ("is_interrupted", "是否中断", "Event", "Events", "1 if pred_segments>=2 else 0", "中断判定核心"),
        ("interruption_count", "中断次数", "Event/Summary", "Events,Summary", "max(0, pred_segments-1)", "Summary 为聚合和"),
        ("interruption_gaps_sec", "中断间隙列表", "Event", "Events", "相邻重叠段 gap 列表，以';'分隔", "仅中断TP非空"),
        ("pred_active_duration_sec", "预测有效激活总时长", "Event", "Events", "sum(各预测片段时长)", "不含间隙"),
        ("pred_span_duration_sec", "预测首末跨度时长", "Event", "Events", "pred_last_end_abs - pred_first_start_abs", "含间隙"),
        ("double_warning_duration_sec", "双告警时长", "Event", "Events", "pred_span_duration_sec - pred_active_duration_sec", "仅中断TP通常>0"),
        ("double_warning_count", "双告警事件数", "Summary", "Summary", "count(is_interrupted==1 的 TP)", "当前口径等价 interrupted_tp"),
        ("mean_double_warning_duration_sec", "双告警平均时长", "Summary", "Summary", "mean(double_warning_duration_sec)", "仅统计中断TP"),
        ("start_offset_sec", "起点偏差", "Event", "Events", "pred_first_start_abs - gt_start_abs", "TP 事件"),
        ("end_offset_sec", "终点偏差", "Event", "Events", "pred_last_end_abs - gt_end_abs", "TP 事件"),
        ("ea_duration_sec", "提前激活时长", "Event", "Events", "max(0, -start_offset_sec)", "TP 事件"),
        ("ed_duration_sec", "提前消失时长", "Event", "Events", "max(0, -end_offset_sec)", "TP 事件"),
        ("la_duration_sec", "延迟激活时长", "Event", "Events", "max(0, start_offset_sec)", "TP 事件"),
        ("ld_duration_sec", "延迟消失时长", "Event", "Events", "max(0, end_offset_sec)", "TP 事件"),
        ("is_ea", "是否提前激活", "Event", "Events", "1 if ea_duration_sec > ea_threshold_sec else 0", "阈值可配置"),
        ("is_ed", "是否提前消失", "Event", "Events", "1 if ed_duration_sec > ed_threshold_sec else 0", "阈值可配置"),
        ("is_la", "是否延迟激活", "Event", "Events", "1 if la_duration_sec > la_threshold_sec else 0", "阈值可配置"),
        ("is_ld", "是否延迟消失", "Event", "Events", "1 if ld_duration_sec > ld_threshold_sec else 0", "阈值可配置"),
        ("kpi_tolerance_sec", "KPI容差秒", "Summary", "Summary", "当前 strict-overlap 口径固定为0", "--tol仅兼容参数"),
        ("ea_threshold_sec/ed_threshold_sec/la_threshold_sec/ld_threshold_sec", "EA/ED/LA/LD阈值", "Summary", "Summary", "来自运行参数或默认值0.25", "会写入 summary"),
    ]
    for row in metric_rows:
        ws_metric_dict.append(list(row))

    ws_tpl_main = wb.create_sheet("Template5G_Main")
    tpl_main_headers = [
        "SourceBagName(来源bag名)",
        "SourceBagPath(来源bag路径)",
        "ReactionType(反应类型)",
        "RadarDirection(雷达方向)",
        "GT_Count(GT区间数)",
        "Pred_Count(预测区间数)",
        "TP(真阳性)",
        "FN(漏报)",
        "FP(误报)",
        "Recall_TPR(召回率)",
        "FPR(误报率)",
        "Precision(精确率)",
        "F1(F1分数)",
        "MeanTPDurSec(TP平均时长秒)",
        "MeanFNDurSec(FN平均时长秒)",
        "MeanFPDurSec(FP平均时长秒)",
        "EACount(提前激活数量)",
        "EDCount(提前消失数量)",
        "LACount(延迟激活数量)",
        "LDCount(延迟消失数量)",
        "DoubleWarnCount(双告警数量)",
        "ManualConclusion(人工结论)",
        "ManualComment(人工备注)",
        "PER(算法评审)",
        "SIT(系统评审)",
        "FCT(功能评审)",
        "Category(分类)",
        "RecallPct(召回率百分比)",
    ]
    if not has_source_bag_cols_summary:
        tpl_main_headers = tpl_main_headers[2:]
    ws_tpl_main.append(tpl_main_headers)
    recall_col_idx = tpl_main_headers.index("Recall_TPR(召回率)") + 1
    try:
        from openpyxl.utils import get_column_letter
        recall_col_letter = get_column_letter(recall_col_idx)
    except Exception:
        recall_col_letter = "H"
    row_idx = 2
    for r in summary_rows:
        row_vals = [
                r.get("source_bag_name", ""),
                r.get("source_bag_path", ""),
                r.get("label_name", ""),
                r.get("radar_direction", ""),
                r.get("gt_intervals", ""),
                r.get("pred_intervals", ""),
                r.get("tp", ""),
                r.get("fn", ""),
                r.get("fp", ""),
                r.get("tpr", ""),
                r.get("fpr", ""),
                r.get("precision", ""),
                r.get("f1", ""),
                r.get("mean_tp_duration_sec", ""),
                r.get("mean_fn_duration_sec", ""),
                r.get("mean_fp_duration_sec", ""),
                r.get("ea_count", ""),
                r.get("ed_count", ""),
                r.get("la_count", ""),
                r.get("ld_count", ""),
                r.get("double_warning_count", ""),
                "",
                "",
                "",
                "",
                "",
                "",
                f"=IFERROR({recall_col_letter}{row_idx}*100,\"\")",
            ]
        if not has_source_bag_cols_summary:
            row_vals = row_vals[2:]
        ws_tpl_main.append(row_vals)
        row_idx += 1

    ws_tpl_events = wb.create_sheet("Template5G_Events")
    tpl_event_headers = [
        "SourceBagName(来源bag名)",
        "SourceBagPath(来源bag路径)",
        "ReactionType(反应类型)",
        "RadarDirection(雷达方向)",
        "EventType(事件类型)",
        "GT_Start_Abs(真值开始绝对秒)",
        "GT_Start_CN(真值开始北京时间)",
        "GT_End_Abs(真值结束绝对秒)",
        "GT_End_CN(真值结束北京时间)",
        "Pred_Start_Abs(预测开始绝对秒)",
        "Pred_Start_CN(预测开始北京时间)",
        "Pred_End_Abs(预测结束绝对秒)",
        "Pred_End_CN(预测结束北京时间)",
        "DelaySec(延迟秒)",
        "OverlapSec(重叠秒)",
        "GT_DurationSec(真值时长秒)",
        "Pred_Active_DurationSec(预测有效时长秒)",
        "Pred_Segments(预测分段数)",
        "IsInterrupted(是否中断)",
        "IsEA(是否提前激活)",
        "IsED(是否提前消失)",
        "IsLA(是否延迟激活)",
        "IsLD(是否延迟消失)",
        "ManualLabeling(人工标注结论)",
        "ManualComment(人工备注)",
        "PER(算法评审)",
        "SIT(系统评审)",
        "FCT(功能评审)",
        "Category(分类)",
        "NotConsidered(是否不纳统)",
        "NotConsideredReason(不纳统原因)",
    ]
    if not has_source_bag_cols:
        tpl_event_headers = tpl_event_headers[2:]
    ws_tpl_events.append(tpl_event_headers)
    for r in event_rows:
        row_vals = [
                r.get("source_bag_name", ""),
                r.get("source_bag_path", ""),
                r.get("label_name", ""),
                r.get("radar_direction", ""),
                r.get("event_type", ""),
                r.get("gt_start_abs", ""),
                ros_abs_sec_to_cn_text(r.get("gt_start_abs", "")),
                r.get("gt_end_abs", ""),
                ros_abs_sec_to_cn_text(r.get("gt_end_abs", "")),
                r.get("pred_start_abs", ""),
                ros_abs_sec_to_cn_text(r.get("pred_start_abs", "")),
                r.get("pred_end_abs", ""),
                ros_abs_sec_to_cn_text(r.get("pred_end_abs", "")),
                r.get("delay_sec", ""),
                r.get("overlap_sec", ""),
                r.get("gt_duration_sec", ""),
                r.get("pred_active_duration_sec", ""),
                r.get("pred_segments", ""),
                r.get("is_interrupted", ""),
                r.get("is_ea", ""),
                r.get("is_ed", ""),
                r.get("is_la", ""),
                r.get("is_ld", ""),
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        if not has_source_bag_cols:
            row_vals = row_vals[2:]
        ws_tpl_events.append(row_vals)

    ws_tpl_guide = wb.create_sheet("Template5G_ManualGuide")
    ws_tpl_guide.append(["Field(字段)", "Type(类型)", "Meaning(含义)", "Example(示例)"])
    manual_guides = [
        ("ManualConclusion(人工结论)", "Manual(人工填写)", "对该行统计结论的最终判断", "通过/关注"),
        ("ManualComment(人工备注)", "Manual(人工填写)", "补充说明异常原因或场景信息", "雨天反光导致误报"),
        ("PER(算法评审)", "Manual(人工填写)", "算法侧评审结论", "确认算法策略问题"),
        ("SIT(系统评审)", "Manual(人工填写)", "系统集成测试评审", "传感器时序正常"),
        ("FCT(功能评审)", "Manual(人工填写)", "功能测试评审", "场景符合测试规范"),
        ("Category(分类)", "Manual(人工填写)", "问题分类标签", "时序/阈值/标注"),
        ("NotConsidered(是否不纳统)", "Manual(人工填写)", "该事件是否不计入 KPI", "0/1"),
        ("NotConsideredReason(不纳统原因)", "Manual(人工填写)", "若不纳统，填写原因", "施工路段不在设计域"),
    ]
    for row in manual_guides:
        ws_tpl_guide.append(list(row))

    ws_meta = wb.create_sheet("Meta")
    ws_meta.append(["key", "value"])
    ws_meta.append(["summary_csv_path", summary_csv_path])
    ws_meta.append(["events_csv_path", events_csv_path])
    ws_meta.append(["generated_at_utc", summary_rows[0].get("generated_at_utc", "") if summary_rows else ""])
    ws_meta.append(["generated_at_cn", summary_rows[0].get("generated_at_cn", "") if summary_rows else ""])
    ws_meta.append(["abs_time_human_timezone", "Asia/Shanghai (UTC+08:00)"])
    ws_meta.append(["abs_time_human_format", "YYYY-MM-DD HH:MM:SS.mmm"])
    if report_note:
        ws_meta.append(["note", report_note])

    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    wb.save(excel_path)


def to_csv_field(s: str) -> str:
    if s is None:
        return ""
    if any(c in s for c in [",", '"', "\n", "\r"]):
        return '"' + s.replace('"', '""') + '"'
    return s


def find_col(header_norm: List[str], *cands: str) -> int:
    for c in cands:
        c_norm = c.strip().lower()
        for i, h in enumerate(header_norm):
            if h == c_norm:
                return i
    return -1


def load_gt_csv(csv_path: str) -> Tuple[List[Interval], CsvMeta, List[str]]:
    warnings: List[str] = []

    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        try:
            header = next(reader)
        except StopIteration:
            raise ValueError(f"CSV is empty: {csv_path}")

        header_norm = [trim(h).lower() for h in header]
        radar_col = find_col(header_norm, "camera_radar_id", "radar_id")
        label_col = find_col(header_norm, "label_name", "label")
        start_col = find_col(header_norm, "start_time_abs", "start_time")
        end_col = find_col(header_norm, "end_time_abs", "end_time")

        bag_start_col = find_col(header_norm, "bag_start_time")
        bag_end_col = find_col(header_norm, "bag_end_time")
        bag_path_col = find_col(header_norm, "bag_path")
        bag_name_col = find_col(header_norm, "bag_name")

        if radar_col < 0 or label_col < 0 or start_col < 0 or end_col < 0:
            raise ValueError(
                "CSV missing required columns. Need at least: "
                "camera_radar_id/radar_id, label_name/label, start_time_abs/start_time, end_time_abs/end_time"
            )

        loaded: List[Interval] = []
        bag_start_sec = 0.0
        bag_end_sec = 0.0
        bag_path = ""
        bag_name = ""

        invalid_radar_count = 0
        unknown_label_count = 0
        malformed_count = 0

        for row in reader:
            if not row or all(not trim(x) for x in row):
                continue

            def cell(idx: int) -> str:
                if idx < 0 or idx >= len(row):
                    return ""
                return trim(row[idx])

            rid = parse_int(cell(radar_col))
            if rid is None:
                malformed_count += 1
                continue
            if rid < 1 or rid > K_MAX_RADAR_ID:
                invalid_radar_count += 1
                continue

            label_name_in = cell(label_col)
            label_idx = adas_label_index_from_name(label_name_in)
            if label_idx < 1 or label_idx > K_ADAS_WARN_COUNT:
                unknown_label_count += 1
                continue

            start_sec = parse_float(cell(start_col))
            end_sec = parse_float(cell(end_col))
            if start_sec is None or end_sec is None:
                malformed_count += 1
                continue
            if end_sec < start_sec:
                start_sec, end_sec = end_sec, start_sec
            if abs(end_sec - start_sec) < 1e-9:
                end_sec = start_sec + 1e-3

            if bag_start_col >= 0:
                v = parse_float(cell(bag_start_col))
                if v is not None and v > 0.0 and (bag_start_sec <= 0.0 or v < bag_start_sec):
                    bag_start_sec = v
            if bag_end_col >= 0:
                v = parse_float(cell(bag_end_col))
                if v is not None and v > 0.0 and v > bag_end_sec:
                    bag_end_sec = v
            if not bag_path and bag_path_col >= 0:
                bag_path = cell(bag_path_col)
            if not bag_name and bag_name_col >= 0:
                bag_name = cell(bag_name_col)

            loaded.append(
                Interval(
                    radar_id=rid,
                    label_index=label_idx,
                    label_name=K_ADAS_LABEL_NAMES[label_idx],
                    start_sec=float(start_sec),
                    end_sec=float(end_sec),
                )
            )

    if not loaded:
        reason = [
            "No valid GT intervals found in CSV.",
            f"stats: invalid_radar_id={invalid_radar_count}, unknown_label={unknown_label_count}, malformed={malformed_count}",
        ]
        if invalid_radar_count > 0:
            reason.append("hint: camera_radar_id may be 0 (UNASSIGNED mapping).")
        if unknown_label_count > 0:
            reason.append("hint: label_name not in 15 ADAS labels (BSD/LCA/DOW/RCW/RCTA/RCTB/FCTA/FCTB).")
        if malformed_count > 0:
            reason.append("hint: start_time_abs/end_time_abs may be non-numeric.")
        raise ValueError("\n".join(reason))

    loaded.sort(key=lambda x: (x.start_sec, x.end_sec, x.radar_id, x.label_index))

    if bag_start_sec <= 0.0:
        bag_start_sec = loaded[0].start_sec
    if bag_end_sec <= bag_start_sec:
        bag_end_sec = loaded[-1].end_sec
    if not bag_name and bag_path:
        bag_name = os.path.basename(bag_path)

    if unknown_label_count > 0:
        warnings.append(f"ignored unknown labels: {unknown_label_count}")
    if invalid_radar_count > 0:
        warnings.append(f"ignored invalid radar_id rows: {invalid_radar_count}")
    if malformed_count > 0:
        warnings.append(f"ignored malformed rows: {malformed_count}")

    meta = CsvMeta(
        csv_path=os.path.abspath(csv_path),
        bag_path=bag_path,
        bag_name=bag_name,
        bag_start_sec=bag_start_sec,
        bag_end_sec=bag_end_sec,
    )
    return loaded, meta, warnings


def interval_overlap_sec(s0: float, e0: float, s1: float, e1: float) -> float:
    return max(0.0, min(e0, e1) - max(s0, s1))


def is_interval_match_strict_overlap(s0: float, e0: float, s1: float, e1: float) -> bool:
    return interval_overlap_sec(s0, e0, s1, e1) > 0.0


def _merge_segments(segments: List[Tuple[float, float]], merge_tol: float = 1e-9) -> List[Tuple[float, float]]:
    if not segments:
        return []
    segments = sorted(segments, key=lambda x: (x[0], x[1]))
    merged: List[Tuple[float, float]] = [segments[0]]
    for s, e in segments[1:]:
        last_s, last_e = merged[-1]
        if s <= last_e + merge_tol:
            merged[-1] = (last_s, max(last_e, e))
        else:
            merged.append((s, e))
    return merged


def resolve_event_sec(
    radar_id: int,
    last_lgu_sec: List[float],
    bag_now_sec: float,
) -> float:
    lgu_sec = 0.0
    if 1 <= radar_id <= K_MAX_RADAR_ID:
        lgu_sec = last_lgu_sec[radar_id]
    return lgu_sec if lgu_sec > 0.0 else bag_now_sec


def build_pred_intervals(
    bag_path: str,
    warning_topic: str,
    lgu_prefix: str,
) -> Tuple[List[Interval], Dict[str, float]]:
    stats: Dict[str, float] = {
        "warning_msgs": 0,
        "warning_short_msgs": 0,
        "warning_invalid_radar": 0,
        "lgu_msgs": 0,
        "lgu_valid_stamp_msgs": 0,
    }

    topic_to_rid = {f"{lgu_prefix}{i}": i for i in range(1, K_MAX_RADAR_ID + 1)}
    read_topics = [warning_topic] + list(topic_to_rid.keys())

    pred: List[Interval] = []
    active = [[False] * (K_ADAS_WARN_COUNT + 1) for _ in range(K_MAX_RADAR_ID + 1)]
    start_sec = [[0.0] * (K_ADAS_WARN_COUNT + 1) for _ in range(K_MAX_RADAR_ID + 1)]
    last_lgu_sec = [0.0] * (K_MAX_RADAR_ID + 1)
    last_warn_sec = [0.0] * (K_MAX_RADAR_ID + 1)

    with rosbag.Bag(bag_path, "r") as bag:
        bag_end = bag.get_end_time()
        for topic, msg, bag_t in bag.read_messages(topics=read_topics):
            bag_now = float(bag_t.to_sec())

            if topic in topic_to_rid:
                stats["lgu_msgs"] += 1
                rid = topic_to_rid[topic]
                ts = 0.0
                try:
                    ts = float(msg.header.stamp.toSec())
                except Exception:
                    ts = 0.0
                if ts <= 0.0:
                    ts = bag_now
                if ts > 0.0:
                    stats["lgu_valid_stamp_msgs"] += 1
                    last_lgu_sec[rid] = ts
                continue

            if topic != warning_topic:
                continue

            stats["warning_msgs"] += 1
            data = list(getattr(msg, "data", []))
            if len(data) < K_ADAS_WARN_COUNT + 1:
                stats["warning_short_msgs"] += 1
                continue

            rid = int(data[0])
            if rid < 1 or rid > K_MAX_RADAR_ID:
                stats["warning_invalid_radar"] += 1
                continue

            event_sec = resolve_event_sec(rid, last_lgu_sec, bag_now)

            # Handle backward time jump (seek/restart): clear active states and pred intervals to avoid timeline mix.
            if last_warn_sec[rid] > 0.0 and event_sec + 0.5 < last_warn_sec[rid]:
                pred.clear()
                for rr in range(1, K_MAX_RADAR_ID + 1):
                    for li in range(1, K_ADAS_WARN_COUNT + 1):
                        active[rr][li] = False
                        start_sec[rr][li] = 0.0
                    last_warn_sec[rr] = 0.0

            last_warn_sec[rid] = event_sec

            for li in range(1, K_ADAS_WARN_COUNT + 1):
                now_on = int(data[li]) > 0
                prev_on = active[rid][li]

                if now_on and not prev_on:
                    active[rid][li] = True
                    start_sec[rid][li] = event_sec
                    continue

                if (not now_on) and prev_on:
                    s = start_sec[rid][li]
                    if s <= 0.0 or event_sec < s:
                        s = event_sec
                    pred.append(
                        Interval(
                            radar_id=rid,
                            label_index=li,
                            label_name=K_ADAS_LABEL_NAMES[li],
                            start_sec=s,
                            end_sec=event_sec,
                        )
                    )
                    active[rid][li] = False
                    start_sec[rid][li] = 0.0

        # Close remaining active intervals at end of bag.
        for rid in range(1, K_MAX_RADAR_ID + 1):
            for li in range(1, K_ADAS_WARN_COUNT + 1):
                if not active[rid][li]:
                    continue
                s = start_sec[rid][li]
                e = last_warn_sec[rid]
                if e <= 0.0:
                    e = last_lgu_sec[rid]
                if e <= 0.0:
                    e = bag_end if bag_end > 0.0 else s
                if e < s:
                    e = s
                pred.append(
                    Interval(
                        radar_id=rid,
                        label_index=li,
                        label_name=K_ADAS_LABEL_NAMES[li],
                        start_sec=s,
                        end_sec=e,
                    )
                )

    pred.sort(key=lambda x: (x.start_sec, x.end_sec, x.radar_id, x.label_index))
    return pred, stats


def compute_metrics(
    gt: List[Interval],
    pred: List[Interval],
    tol_sec: float,
    ea_threshold_sec: float,
    ed_threshold_sec: float,
    la_threshold_sec: float,
    ld_threshold_sec: float,
) -> Tuple[List[MetricRow], List[EventRow]]:
    _ = tol_sec  # kept for CLI compatibility; PM standard uses strict overlap without tolerance.
    gt_by: Dict[Tuple[int, int], List[Interval]] = {}
    pred_by: Dict[Tuple[int, int], List[Interval]] = {}

    for it in gt:
        gt_by.setdefault((it.radar_id, it.label_index), []).append(it)
    for it in pred:
        pred_by.setdefault((it.radar_id, it.label_index), []).append(it)

    keys = sorted(set(gt_by.keys()) | set(pred_by.keys()))

    metrics: List[MetricRow] = []
    events: List[EventRow] = []

    total_gt = total_pred = total_tp = total_fn = total_fp = 0
    total_interrupted_tp = 0
    total_interruption_count = 0
    total_delays: List[float] = []
    total_overlaps: List[float] = []
    total_interrupt_gaps: List[float] = []
    total_tp_durations: List[float] = []
    total_fn_durations: List[float] = []
    total_fp_durations: List[float] = []
    total_ea_durations: List[float] = []
    total_ed_durations: List[float] = []
    total_la_durations: List[float] = []
    total_ld_durations: List[float] = []
    total_double_warning_durations: List[float] = []

    for key in keys:
        g = sorted(gt_by.get(key, []), key=lambda x: (x.start_sec, x.end_sec))
        p = sorted(pred_by.get(key, []), key=lambda x: (x.start_sec, x.end_sec))

        pred_overlaps_any = [False] * len(p)
        tp = fn = fp = 0
        interrupted_tp = 0
        interruption_count = 0
        delays: List[float] = []
        overlaps: List[float] = []
        interrupt_gaps: List[float] = []
        tp_durations: List[float] = []
        fn_durations: List[float] = []
        fp_durations: List[float] = []
        ea_durations: List[float] = []
        ed_durations: List[float] = []
        la_durations: List[float] = []
        ld_durations: List[float] = []
        double_warning_durations: List[float] = []

        rid, li = key
        lname = K_ADAS_LABEL_NAMES.get(li, "")

        for gi, gv in enumerate(g):
            overlap_idx: List[int] = []
            for pi, pv in enumerate(p):
                if is_interval_match_strict_overlap(gv.start_sec, gv.end_sec, pv.start_sec, pv.end_sec):
                    overlap_idx.append(pi)
                    pred_overlaps_any[pi] = True

            if not overlap_idx:
                fn += 1
                gt_duration = max(0.0, gv.end_sec - gv.start_sec)
                fn_durations.append(gt_duration)
                total_fn_durations.append(gt_duration)
                events.append(
                    EventRow(
                        radar_id=rid,
                        label_index=li,
                        label_name=lname,
                        event_type="FN",
                        gt_start=gv.start_sec,
                        gt_end=gv.end_sec,
                        pred_start=math.nan,
                        pred_end=math.nan,
                        pred_first_start=math.nan,
                        pred_last_end=math.nan,
                        pred_active_duration=math.nan,
                        pred_span_duration=math.nan,
                        delay=math.nan,
                        overlap=math.nan,
                        pred_segments=0,
                        interruption_count=0,
                        interruption_gaps=[],
                        is_interrupted=0,
                        start_offset=math.nan,
                        end_offset=math.nan,
                        ea_duration=math.nan,
                        ed_duration=math.nan,
                        la_duration=math.nan,
                        ld_duration=math.nan,
                        is_ea=0,
                        is_ed=0,
                        is_la=0,
                        is_ld=0,
                        double_warning_duration=math.nan,
                    )
                )
                continue

            tp += 1
            overlap_idx.sort(key=lambda pi: (p[pi].start_sec, p[pi].end_sec))
            primary = p[overlap_idx[0]]
            first_pred_start = primary.start_sec
            last_pred_end = max(p[pi].end_sec for pi in overlap_idx)
            delay = primary.start_sec - gv.start_sec
            delays.append(delay)
            total_delays.append(delay)

            clipped: List[Tuple[float, float]] = []
            for pi in overlap_idx:
                s = max(gv.start_sec, p[pi].start_sec)
                e = min(gv.end_sec, p[pi].end_sec)
                if e >= s:
                    clipped.append((s, e))
            merged = _merge_segments(clipped)
            total_overlap = sum(max(0.0, e - s) for s, e in merged)
            overlaps.append(total_overlap)
            total_overlaps.append(total_overlap)

            full_segments = _merge_segments([(p[pi].start_sec, p[pi].end_sec) for pi in overlap_idx])
            pred_active_duration = sum(max(0.0, e - s) for s, e in full_segments)
            pred_span_duration = max(0.0, last_pred_end - first_pred_start)
            tp_durations.append(pred_active_duration)
            total_tp_durations.append(pred_active_duration)

            start_offset = first_pred_start - gv.start_sec
            end_offset = last_pred_end - gv.end_sec
            ea_duration = max(0.0, -start_offset)
            ed_duration = max(0.0, -end_offset)
            la_duration = max(0.0, start_offset)
            ld_duration = max(0.0, end_offset)
            is_ea = 1 if ea_duration > ea_threshold_sec else 0
            is_ed = 1 if ed_duration > ed_threshold_sec else 0
            is_la = 1 if la_duration > la_threshold_sec else 0
            is_ld = 1 if ld_duration > ld_threshold_sec else 0
            if is_ea:
                ea_durations.append(ea_duration)
                total_ea_durations.append(ea_duration)
            if is_ed:
                ed_durations.append(ed_duration)
                total_ed_durations.append(ed_duration)
            if is_la:
                la_durations.append(la_duration)
                total_la_durations.append(la_duration)
            if is_ld:
                ld_durations.append(ld_duration)
                total_ld_durations.append(ld_duration)

            pred_segments = len(merged)
            gt_interruptions = max(0, pred_segments - 1)
            gaps: List[float] = []
            double_warning_duration = max(0.0, pred_span_duration - pred_active_duration)
            if gt_interruptions > 0:
                interrupted_tp += 1
                interruption_count += gt_interruptions
                double_warning_durations.append(double_warning_duration)
                total_double_warning_durations.append(double_warning_duration)
                for i in range(len(merged) - 1):
                    gap = max(0.0, merged[i + 1][0] - merged[i][1])
                    gaps.append(gap)
                    interrupt_gaps.append(gap)
                    total_interrupt_gaps.append(gap)

            events.append(
                EventRow(
                    radar_id=rid,
                    label_index=li,
                    label_name=lname,
                    event_type="TP",
                    gt_start=gv.start_sec,
                    gt_end=gv.end_sec,
                    pred_start=primary.start_sec,
                    pred_end=primary.end_sec,
                    pred_first_start=first_pred_start,
                    pred_last_end=last_pred_end,
                    pred_active_duration=pred_active_duration,
                    pred_span_duration=pred_span_duration,
                    delay=delay,
                    overlap=total_overlap,
                    pred_segments=pred_segments,
                    interruption_count=gt_interruptions,
                    interruption_gaps=gaps,
                    is_interrupted=1 if gt_interruptions > 0 else 0,
                    start_offset=start_offset,
                    end_offset=end_offset,
                    ea_duration=ea_duration,
                    ed_duration=ed_duration,
                    la_duration=la_duration,
                    ld_duration=ld_duration,
                    is_ea=is_ea,
                    is_ed=is_ed,
                    is_la=is_la,
                    is_ld=is_ld,
                    double_warning_duration=double_warning_duration,
                )
            )

        for pi, pv in enumerate(p):
            if pred_overlaps_any[pi]:
                continue
            fp += 1
            fp_duration = max(0.0, pv.end_sec - pv.start_sec)
            fp_durations.append(fp_duration)
            total_fp_durations.append(fp_duration)
            events.append(
                EventRow(
                    radar_id=rid,
                    label_index=li,
                    label_name=lname,
                    event_type="FP",
                    gt_start=math.nan,
                    gt_end=math.nan,
                    pred_start=pv.start_sec,
                    pred_end=pv.end_sec,
                    pred_first_start=pv.start_sec,
                    pred_last_end=pv.end_sec,
                    pred_active_duration=max(0.0, pv.end_sec - pv.start_sec),
                    pred_span_duration=max(0.0, pv.end_sec - pv.start_sec),
                    delay=math.nan,
                    overlap=math.nan,
                    pred_segments=1,
                    interruption_count=0,
                    interruption_gaps=[],
                    is_interrupted=0,
                    start_offset=math.nan,
                    end_offset=math.nan,
                    ea_duration=math.nan,
                    ed_duration=math.nan,
                    la_duration=math.nan,
                    ld_duration=math.nan,
                    is_ea=0,
                    is_ed=0,
                    is_la=0,
                    is_ld=0,
                    double_warning_duration=math.nan,
                )
            )

        denom_tp_fn = tp + fn
        denom_all = tp + fn + fp
        denom_prec = tp + fp
        tpr = (tp / denom_tp_fn) if denom_tp_fn > 0 else 0.0
        fnr = (fn / denom_tp_fn) if denom_tp_fn > 0 else 0.0
        fpr = (fp / denom_all) if denom_all > 0 else 0.0
        precision = (tp / denom_prec) if denom_prec > 0 else 0.0
        f1 = (2.0 * precision * tpr / (precision + tpr)) if (precision + tpr) > 0 else 0.0

        mean_delay = min_delay = max_delay = mean_overlap = math.nan
        if delays:
            mean_delay = sum(delays) / len(delays)
            min_delay = min(delays)
            max_delay = max(delays)
        if overlaps:
            mean_overlap = sum(overlaps) / len(overlaps)

        mean_interrupt_gap = max_interrupt_gap = math.nan
        if interrupt_gaps:
            mean_interrupt_gap = sum(interrupt_gaps) / len(interrupt_gaps)
            max_interrupt_gap = max(interrupt_gaps)

        interruption_tp_ratio = (interrupted_tp / tp) if tp > 0 else 0.0
        mean_tp_duration = mean_or_nan(tp_durations)
        mean_fn_duration = mean_or_nan(fn_durations)
        mean_fp_duration = mean_or_nan(fp_durations)
        mean_ea_duration = mean_or_nan(ea_durations)
        mean_ed_duration = mean_or_nan(ed_durations)
        mean_la_duration = mean_or_nan(la_durations)
        mean_ld_duration = mean_or_nan(ld_durations)
        mean_double_warning_duration = mean_or_nan(double_warning_durations)

        metrics.append(
            MetricRow(
                radar_id=rid,
                label_index=li,
                label_name=lname,
                gt_count=len(g),
                pred_count=len(p),
                tp=tp,
                fn=fn,
                fp=fp,
                tpr=tpr,
                fpr=fpr,
                fnr=fnr,
                precision=precision,
                f1=f1,
                mean_delay=mean_delay,
                min_delay=min_delay,
                max_delay=max_delay,
                mean_overlap=mean_overlap,
                interrupted_tp=interrupted_tp,
                interruption_count=interruption_count,
                interruption_tp_ratio=interruption_tp_ratio,
                mean_interrupt_gap=mean_interrupt_gap,
                max_interrupt_gap=max_interrupt_gap,
                mean_tp_duration=mean_tp_duration,
                mean_fn_duration=mean_fn_duration,
                mean_fp_duration=mean_fp_duration,
                ea_count=len(ea_durations),
                mean_ea_duration=mean_ea_duration,
                ed_count=len(ed_durations),
                mean_ed_duration=mean_ed_duration,
                la_count=len(la_durations),
                mean_la_duration=mean_la_duration,
                ld_count=len(ld_durations),
                mean_ld_duration=mean_ld_duration,
                double_warning_count=interrupted_tp,
                mean_double_warning_duration=mean_double_warning_duration,
            )
        )

        total_gt += len(g)
        total_pred += len(p)
        total_tp += tp
        total_fn += fn
        total_fp += fp
        total_interrupted_tp += interrupted_tp
        total_interruption_count += interruption_count

    denom_tp_fn = total_tp + total_fn
    denom_all = total_tp + total_fn + total_fp
    denom_prec = total_tp + total_fp
    total_tpr = (total_tp / denom_tp_fn) if denom_tp_fn > 0 else 0.0
    total_fnr = (total_fn / denom_tp_fn) if denom_tp_fn > 0 else 0.0
    total_fpr = (total_fp / denom_all) if denom_all > 0 else 0.0
    total_precision = (total_tp / denom_prec) if denom_prec > 0 else 0.0
    total_f1 = (2.0 * total_precision * total_tpr / (total_precision + total_tpr)) if (total_precision + total_tpr) > 0 else 0.0
    total_interruption_tp_ratio = (total_interrupted_tp / total_tp) if total_tp > 0 else 0.0

    total_mean_delay = total_min_delay = total_max_delay = total_mean_overlap = math.nan
    if total_delays:
        total_mean_delay = sum(total_delays) / len(total_delays)
        total_min_delay = min(total_delays)
        total_max_delay = max(total_delays)
    if total_overlaps:
        total_mean_overlap = sum(total_overlaps) / len(total_overlaps)

    total_mean_interrupt_gap = total_max_interrupt_gap = math.nan
    if total_interrupt_gaps:
        total_mean_interrupt_gap = sum(total_interrupt_gaps) / len(total_interrupt_gaps)
        total_max_interrupt_gap = max(total_interrupt_gaps)
    total_mean_tp_duration = mean_or_nan(total_tp_durations)
    total_mean_fn_duration = mean_or_nan(total_fn_durations)
    total_mean_fp_duration = mean_or_nan(total_fp_durations)
    total_mean_ea_duration = mean_or_nan(total_ea_durations)
    total_mean_ed_duration = mean_or_nan(total_ed_durations)
    total_mean_la_duration = mean_or_nan(total_la_durations)
    total_mean_ld_duration = mean_or_nan(total_ld_durations)
    total_mean_double_warning_duration = mean_or_nan(total_double_warning_durations)

    metrics.append(
        MetricRow(
            radar_id=0,
            label_index=0,
            label_name="ALL",
            gt_count=total_gt,
            pred_count=total_pred,
            tp=total_tp,
            fn=total_fn,
            fp=total_fp,
            tpr=total_tpr,
            fpr=total_fpr,
            fnr=total_fnr,
            precision=total_precision,
            f1=total_f1,
            mean_delay=total_mean_delay,
            min_delay=total_min_delay,
            max_delay=total_max_delay,
            mean_overlap=total_mean_overlap,
            interrupted_tp=total_interrupted_tp,
            interruption_count=total_interruption_count,
            interruption_tp_ratio=total_interruption_tp_ratio,
            mean_interrupt_gap=total_mean_interrupt_gap,
            max_interrupt_gap=total_max_interrupt_gap,
            mean_tp_duration=total_mean_tp_duration,
            mean_fn_duration=total_mean_fn_duration,
            mean_fp_duration=total_mean_fp_duration,
            ea_count=len(total_ea_durations),
            mean_ea_duration=total_mean_ea_duration,
            ed_count=len(total_ed_durations),
            mean_ed_duration=total_mean_ed_duration,
            la_count=len(total_la_durations),
            mean_la_duration=total_mean_la_duration,
            ld_count=len(total_ld_durations),
            mean_ld_duration=total_mean_ld_duration,
            double_warning_count=total_interrupted_tp,
            mean_double_warning_duration=total_mean_double_warning_duration,
        )
    )

    return metrics, events


def export_kpi(
    summary_path: str,
    events_path: str,
    metrics: List[MetricRow],
    events: List[EventRow],
    tol_sec: float,
    ea_threshold_sec: float,
    ed_threshold_sec: float,
    la_threshold_sec: float,
    ld_threshold_sec: float,
    csv_meta: CsvMeta,
    playback_bag_path: str,
) -> None:
    _ = tol_sec  # kept for compatibility; summary writes 0.0 in PM strict-overlap mode.
    now_utc_dt = dt.datetime.now(dt.timezone.utc)
    now_cn_dt = now_utc_dt.astimezone(dt.timezone(dt.timedelta(hours=8)))
    now_utc = now_utc_dt.strftime("%Y-%m-%dT%H:%M:%SZ")
    now_cn = now_cn_dt.strftime("%Y-%m-%dT%H:%M:%S+08:00")
    csv_path = csv_meta.csv_path
    gt_bag_path = csv_meta.bag_path
    bag_start = csv_meta.bag_start_sec

    os.makedirs(os.path.dirname(summary_path), exist_ok=True)

    with open(summary_path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow([
            "generated_at_utc",
            "generated_at_cn",
            "label_csv_path",
            "playback_bag_path",
            "gt_bag_path",
            "kpi_tolerance_sec",
            "radar_id",
            "radar_direction",
            "label_name",
            "gt_intervals",
            "pred_intervals",
            "tp",
            "fn",
            "fp",
            "tpr",
            "fpr",
            "fnr",
            "precision",
            "f1",
            "mean_delay_sec",
            "min_delay_sec",
            "max_delay_sec",
            "mean_overlap_sec",
            "interrupted_tp",
            "interruption_count",
            "interruption_tp_ratio",
            "mean_interrupt_gap_sec",
            "max_interrupt_gap_sec",
            "mean_tp_duration_sec",
            "mean_fn_duration_sec",
            "mean_fp_duration_sec",
            "ea_count",
            "mean_ea_duration_sec",
            "ed_count",
            "mean_ed_duration_sec",
            "la_count",
            "mean_la_duration_sec",
            "ld_count",
            "mean_ld_duration_sec",
            "double_warning_count",
            "mean_double_warning_duration_sec",
            "ea_threshold_sec",
            "ed_threshold_sec",
            "la_threshold_sec",
            "ld_threshold_sec",
        ])
        for row in metrics:
            has_tp = row.tp > 0
            has_interrupt = row.interruption_count > 0
            w.writerow([
                now_utc,
                now_cn,
                csv_path,
                playback_bag_path,
                gt_bag_path,
                fmt6(0.0),
                row.radar_id,
                radar_direction_from_id(row.radar_id),
                row.label_name,
                row.gt_count,
                row.pred_count,
                row.tp,
                row.fn,
                row.fp,
                fmt6(row.tpr),
                fmt6(row.fpr),
                fmt6(row.fnr),
                fmt6(row.precision),
                fmt6(row.f1),
                (fmt6(row.mean_delay) if has_tp else ""),
                (fmt6(row.min_delay) if has_tp else ""),
                (fmt6(row.max_delay) if has_tp else ""),
                (fmt6(row.mean_overlap) if has_tp else ""),
                row.interrupted_tp,
                row.interruption_count,
                fmt6(row.interruption_tp_ratio),
                (fmt6(row.mean_interrupt_gap) if has_interrupt else ""),
                (fmt6(row.max_interrupt_gap) if has_interrupt else ""),
                (fmt6(row.mean_tp_duration) if row.tp > 0 else ""),
                (fmt6(row.mean_fn_duration) if row.fn > 0 else ""),
                (fmt6(row.mean_fp_duration) if row.fp > 0 else ""),
                row.ea_count,
                (fmt6(row.mean_ea_duration) if row.ea_count > 0 else ""),
                row.ed_count,
                (fmt6(row.mean_ed_duration) if row.ed_count > 0 else ""),
                row.la_count,
                (fmt6(row.mean_la_duration) if row.la_count > 0 else ""),
                row.ld_count,
                (fmt6(row.mean_ld_duration) if row.ld_count > 0 else ""),
                row.double_warning_count,
                (fmt6(row.mean_double_warning_duration) if row.double_warning_count > 0 else ""),
                fmt6(ea_threshold_sec),
                fmt6(ed_threshold_sec),
                fmt6(la_threshold_sec),
                fmt6(ld_threshold_sec),
            ])

    with open(events_path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow([
            "generated_at_utc",
            "generated_at_cn",
            "label_csv_path",
            "playback_bag_path",
            "gt_bag_path",
            "radar_id",
            "radar_direction",
            "label_name",
            "event_type",
            "gt_start_abs",
            "gt_end_abs",
            "pred_start_abs",
            "pred_end_abs",
            "pred_first_start_abs",
            "pred_last_end_abs",
            "gt_start_rel",
            "gt_end_rel",
            "pred_start_rel",
            "pred_end_rel",
            "pred_first_start_rel",
            "pred_last_end_rel",
            "delay_sec",
            "overlap_sec",
            "gt_duration_sec",
            "pred_duration_sec",
            "pred_active_duration_sec",
            "pred_span_duration_sec",
            "pred_segments",
            "interruption_count",
            "interruption_gaps_sec",
            "is_interrupted",
            "start_offset_sec",
            "end_offset_sec",
            "ea_duration_sec",
            "ed_duration_sec",
            "la_duration_sec",
            "ld_duration_sec",
            "is_ea",
            "is_ed",
            "is_la",
            "is_ld",
            "double_warning_duration_sec",
        ])
        for ev in events:
            gt_dur = (max(0.0, ev.gt_end - ev.gt_start) if math.isfinite(ev.gt_start) and math.isfinite(ev.gt_end) else math.nan)
            pred_dur = (max(0.0, ev.pred_end - ev.pred_start) if math.isfinite(ev.pred_start) and math.isfinite(ev.pred_end) else math.nan)
            gt_start_rel = (ev.gt_start - bag_start if math.isfinite(ev.gt_start) and bag_start > 0.0 else math.nan)
            gt_end_rel = (ev.gt_end - bag_start if math.isfinite(ev.gt_end) and bag_start > 0.0 else math.nan)
            pred_start_rel = (ev.pred_start - bag_start if math.isfinite(ev.pred_start) and bag_start > 0.0 else math.nan)
            pred_end_rel = (ev.pred_end - bag_start if math.isfinite(ev.pred_end) and bag_start > 0.0 else math.nan)
            pred_first_start_rel = (ev.pred_first_start - bag_start if math.isfinite(ev.pred_first_start) and bag_start > 0.0 else math.nan)
            pred_last_end_rel = (ev.pred_last_end - bag_start if math.isfinite(ev.pred_last_end) and bag_start > 0.0 else math.nan)
            gaps = ";".join(f"{g:.6f}" for g in ev.interruption_gaps) if ev.interruption_gaps else ""

            w.writerow([
                now_utc,
                now_cn,
                csv_path,
                playback_bag_path,
                gt_bag_path,
                ev.radar_id,
                radar_direction_from_id(ev.radar_id),
                ev.label_name,
                ev.event_type,
                fmt_opt(ev.gt_start),
                fmt_opt(ev.gt_end),
                fmt_opt(ev.pred_start),
                fmt_opt(ev.pred_end),
                fmt_opt(ev.pred_first_start),
                fmt_opt(ev.pred_last_end),
                fmt_opt(gt_start_rel),
                fmt_opt(gt_end_rel),
                fmt_opt(pred_start_rel),
                fmt_opt(pred_end_rel),
                fmt_opt(pred_first_start_rel),
                fmt_opt(pred_last_end_rel),
                fmt_opt(ev.delay),
                fmt_opt(ev.overlap),
                fmt_opt(gt_dur),
                fmt_opt(pred_dur),
                fmt_opt(ev.pred_active_duration),
                fmt_opt(ev.pred_span_duration),
                ev.pred_segments,
                ev.interruption_count,
                gaps,
                ev.is_interrupted,
                fmt_opt(ev.start_offset),
                fmt_opt(ev.end_offset),
                fmt_opt(ev.ea_duration),
                fmt_opt(ev.ed_duration),
                fmt_opt(ev.la_duration),
                fmt_opt(ev.ld_duration),
                ev.is_ea,
                ev.is_ed,
                ev.is_la,
                ev.is_ld,
                fmt_opt(ev.double_warning_duration),
            ])


def collect_jobs(input_dir: str) -> Tuple[List[BatchJob], List[str]]:
    bags: List[str] = []
    csv_map: Dict[str, List[str]] = {}

    for root, _dirs, files in os.walk(input_dir):
        for fn in files:
            lower = fn.lower()
            p = os.path.abspath(os.path.join(root, fn))
            if lower.endswith(".bag"):
                bags.append(p)
            elif lower.endswith(".csv"):
                csv_map.setdefault(fn, []).append(p)

    jobs: List[BatchJob] = []
    missing: List[str] = []

    def pick_same_dir_or_first(cands: List[str], bag_path: str) -> str:
        if not cands:
            return ""
        bag_dir = os.path.dirname(bag_path)
        for c in cands:
            if os.path.dirname(c) == bag_dir:
                return c
        return cands[0]

    for bag in sorted(bags):
        base = os.path.splitext(os.path.basename(bag))[0]
        expect_gt = f"{base}_corner_radar_gt.csv"
        expect_plain = f"{base}.csv"

        csv_path = pick_same_dir_or_first(csv_map.get(expect_gt, []), bag)
        if not csv_path:
            # fallback to base.csv, but avoid camera_mapping files by exact filename only.
            csv_path = pick_same_dir_or_first(csv_map.get(expect_plain, []), bag)

        if not csv_path:
            missing.append(f"{os.path.basename(bag)} (expect: {expect_gt} or {expect_plain})")
            continue

        jobs.append(BatchJob(bag_path=bag, csv_path=csv_path, bag_base=base))

    return jobs, missing


def run_one_job(
    bag_path: str,
    csv_path: str,
    output_dir: str,
    warning_topic: str,
    lgu_prefix: str,
    tol_sec: float,
    ea_threshold_sec: float,
    ed_threshold_sec: float,
    la_threshold_sec: float,
    ld_threshold_sec: float,
    report_format: str,
    verbose: bool,
) -> Tuple[bool, str, str, str, str, str]:
    # returns: ok, detail, summary_path, events_path, xlsx_path, warnings_text
    gt, meta, warns = load_gt_csv(csv_path)

    pred, stats = build_pred_intervals(
        bag_path=bag_path,
        warning_topic=warning_topic,
        lgu_prefix=lgu_prefix,
    )

    metrics, events = compute_metrics(
        gt=gt,
        pred=pred,
        tol_sec=tol_sec,
        ea_threshold_sec=ea_threshold_sec,
        ed_threshold_sec=ed_threshold_sec,
        la_threshold_sec=la_threshold_sec,
        ld_threshold_sec=ld_threshold_sec,
    )

    base = os.path.splitext(os.path.basename(bag_path))[0]
    summary_path = os.path.join(output_dir, f"{base}_adas_kpi_summary.csv")
    events_path = os.path.join(output_dir, f"{base}_adas_kpi_summary_events.csv")
    xlsx_path = os.path.join(output_dir, f"{base}_adas_kpi_report.xlsx")

    export_kpi(
        summary_path=summary_path,
        events_path=events_path,
        metrics=metrics,
        events=events,
        tol_sec=tol_sec,
        ea_threshold_sec=ea_threshold_sec,
        ed_threshold_sec=ed_threshold_sec,
        la_threshold_sec=la_threshold_sec,
        ld_threshold_sec=ld_threshold_sec,
        csv_meta=meta,
        playback_bag_path=os.path.abspath(bag_path),
    )

    report_format = trim(report_format).lower()
    if report_format not in ("csv", "xlsx", "both"):
        raise ValueError(f"Unsupported report_format: {report_format}")

    exported_xlsx = ""
    if report_format in ("xlsx", "both"):
        note = (
            "5G-aligned extra stats included; strict-overlap matching; "
            f"EA/ED/LA/LD thresholds={ea_threshold_sec:.3f}/{ed_threshold_sec:.3f}/{la_threshold_sec:.3f}/{ld_threshold_sec:.3f} sec"
        )
        export_excel_report_from_csv(
            excel_path=xlsx_path,
            summary_csv_path=summary_path,
            events_csv_path=events_path,
            report_note=note,
        )
        exported_xlsx = xlsx_path

    if report_format == "xlsx":
        try:
            os.remove(summary_path)
        except Exception:
            pass
        try:
            os.remove(events_path)
        except Exception:
            pass
        summary_path = ""
        events_path = ""

    details = []
    if warns:
        details.append("; ".join(warns))
    if verbose:
        details.append(
            "warning_msgs={warning_msgs}, lgu_msgs={lgu_msgs}, pred_intervals={pred_count}".format(
                warning_msgs=int(stats.get("warning_msgs", 0)),
                lgu_msgs=int(stats.get("lgu_msgs", 0)),
                pred_count=len(pred),
            )
        )
    if exported_xlsx:
        details.append(f"xlsx={exported_xlsx}")
    detail = "done" if not details else "done; " + " | ".join(details)
    return True, detail, summary_path, events_path, exported_xlsx, "\n".join(warns)


def ensure_output_dir(path: Optional[str], base_dir: str) -> str:
    if path:
        out = os.path.abspath(path)
    else:
        out = os.path.join(base_dir, "kpi_reports_" + dt.datetime.now().strftime("%Y%m%d_%H%M%S"))
    os.makedirs(out, exist_ok=True)
    return out


def write_index(index_path: str, rows: List[Dict[str, str]]) -> None:
    with open(index_path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["bag_path", "csv_path", "status", "summary_path", "events_path", "xlsx_path", "detail"])
        for r in rows:
            w.writerow([
                r.get("bag_path", ""),
                r.get("csv_path", ""),
                r.get("status", ""),
                r.get("summary_path", ""),
                r.get("events_path", ""),
                r.get("xlsx_path", ""),
                r.get("detail", ""),
            ])


def read_summary_rows_from_csv(path: str) -> List[Dict[str, str]]:
    if not path or (not os.path.isfile(path)):
        return []
    with open(path, "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        return [dict(r) for r in reader]


def read_summary_rows_from_xlsx(path: str) -> List[Dict[str, str]]:
    if not path or (not os.path.isfile(path)):
        return []
    try:
        from openpyxl import load_workbook
    except Exception:
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Summary" not in wb.sheetnames:
        return []
    ws = wb["Summary"]
    header: List[str] = []
    rows: List[Dict[str, str]] = []
    for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = [("" if v is None else str(v)) for v in row]
        if ridx == 1:
            header = vals
            continue
        if not header or all(trim(v) == "" for v in vals):
            continue
        item: Dict[str, str] = {}
        for i, h in enumerate(header):
            if not trim(h):
                continue
            item[h] = vals[i] if i < len(vals) else ""
        rows.append(item)
    return rows


def read_sheet_rows_from_xlsx(path: str, sheet_name: str) -> List[Dict[str, str]]:
    if not path or (not os.path.isfile(path)):
        return []
    try:
        from openpyxl import load_workbook
    except Exception:
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    header: List[str] = []
    rows: List[Dict[str, str]] = []
    for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = [("" if v is None else str(v)) for v in row]
        if ridx == 1:
            header = vals
            continue
        if not header or all(trim(v) == "" for v in vals):
            continue
        item: Dict[str, str] = {}
        for i, h in enumerate(header):
            if not trim(h):
                continue
            item[h] = vals[i] if i < len(vals) else ""
        rows.append(item)
    return rows


def pick_all_row(summary_rows: List[Dict[str, str]]) -> Optional[Dict[str, str]]:
    cands = [r for r in summary_rows if is_all_summary_row(r)]
    if not cands:
        return None
    for r in cands:
        rid = str_to_int(r.get("radar_id", ""))
        if rid == 0:
            return r
    return cands[-1]


def build_batch_aggregate_row(rows: List[Dict[str, str]]) -> Dict[str, str]:
    tp = fn = fp = gt = pred = 0
    interrupted_tp = interruption_count = 0
    ea_count = ed_count = la_count = ld_count = 0
    double_warning_count = 0

    delay_weight_sum = delay_weight = 0.0
    overlap_weight_sum = overlap_weight = 0.0
    tp_dur_weight_sum = tp_dur_weight = 0.0
    fn_dur_weight_sum = fn_dur_weight = 0.0
    fp_dur_weight_sum = fp_dur_weight = 0.0
    ea_dur_weight_sum = ea_dur_weight = 0.0
    ed_dur_weight_sum = ed_dur_weight = 0.0
    la_dur_weight_sum = la_dur_weight = 0.0
    ld_dur_weight_sum = ld_dur_weight = 0.0
    dw_dur_weight_sum = dw_dur_weight = 0.0
    int_gap_weight_sum = int_gap_weight = 0.0
    min_delay = math.inf
    max_delay = -math.inf
    max_int_gap = -math.inf

    first_generated_utc = ""
    first_generated_cn = ""
    tol_val = ""
    ea_thr = ed_thr = la_thr = ld_thr = ""

    for r in rows:
        if not first_generated_utc:
            first_generated_utc = trim(r.get("generated_at_utc", ""))
        if not first_generated_cn:
            first_generated_cn = trim(r.get("generated_at_cn", ""))
        if not tol_val:
            tol_val = trim(r.get("kpi_tolerance_sec", ""))
        if not ea_thr:
            ea_thr = trim(r.get("ea_threshold_sec", ""))
        if not ed_thr:
            ed_thr = trim(r.get("ed_threshold_sec", ""))
        if not la_thr:
            la_thr = trim(r.get("la_threshold_sec", ""))
        if not ld_thr:
            ld_thr = trim(r.get("ld_threshold_sec", ""))

        gt_i = str_to_int(r.get("gt_intervals", "0"))
        pred_i = str_to_int(r.get("pred_intervals", "0"))
        tp_i = str_to_int(r.get("tp", "0"))
        fn_i = str_to_int(r.get("fn", "0"))
        fp_i = str_to_int(r.get("fp", "0"))
        int_tp_i = str_to_int(r.get("interrupted_tp", "0"))
        int_cnt_i = str_to_int(r.get("interruption_count", "0"))
        ea_i = str_to_int(r.get("ea_count", "0"))
        ed_i = str_to_int(r.get("ed_count", "0"))
        la_i = str_to_int(r.get("la_count", "0"))
        ld_i = str_to_int(r.get("ld_count", "0"))
        dw_i = str_to_int(r.get("double_warning_count", "0"))

        gt += gt_i
        pred += pred_i
        tp += tp_i
        fn += fn_i
        fp += fp_i
        interrupted_tp += int_tp_i
        interruption_count += int_cnt_i
        ea_count += ea_i
        ed_count += ed_i
        la_count += la_i
        ld_count += ld_i
        double_warning_count += dw_i

        mean_delay = str_to_float(r.get("mean_delay_sec", ""))
        if math.isfinite(mean_delay) and tp_i > 0:
            delay_weight_sum += mean_delay * tp_i
            delay_weight += tp_i

        mean_overlap = str_to_float(r.get("mean_overlap_sec", ""))
        if math.isfinite(mean_overlap) and tp_i > 0:
            overlap_weight_sum += mean_overlap * tp_i
            overlap_weight += tp_i

        min_d = str_to_float(r.get("min_delay_sec", ""))
        max_d = str_to_float(r.get("max_delay_sec", ""))
        if math.isfinite(min_d):
            min_delay = min(min_delay, min_d)
        if math.isfinite(max_d):
            max_delay = max(max_delay, max_d)

        mean_tp_dur = str_to_float(r.get("mean_tp_duration_sec", ""))
        if math.isfinite(mean_tp_dur) and tp_i > 0:
            tp_dur_weight_sum += mean_tp_dur * tp_i
            tp_dur_weight += tp_i
        mean_fn_dur = str_to_float(r.get("mean_fn_duration_sec", ""))
        if math.isfinite(mean_fn_dur) and fn_i > 0:
            fn_dur_weight_sum += mean_fn_dur * fn_i
            fn_dur_weight += fn_i
        mean_fp_dur = str_to_float(r.get("mean_fp_duration_sec", ""))
        if math.isfinite(mean_fp_dur) and fp_i > 0:
            fp_dur_weight_sum += mean_fp_dur * fp_i
            fp_dur_weight += fp_i

        mean_ea_dur = str_to_float(r.get("mean_ea_duration_sec", ""))
        if math.isfinite(mean_ea_dur) and ea_i > 0:
            ea_dur_weight_sum += mean_ea_dur * ea_i
            ea_dur_weight += ea_i
        mean_ed_dur = str_to_float(r.get("mean_ed_duration_sec", ""))
        if math.isfinite(mean_ed_dur) and ed_i > 0:
            ed_dur_weight_sum += mean_ed_dur * ed_i
            ed_dur_weight += ed_i
        mean_la_dur = str_to_float(r.get("mean_la_duration_sec", ""))
        if math.isfinite(mean_la_dur) and la_i > 0:
            la_dur_weight_sum += mean_la_dur * la_i
            la_dur_weight += la_i
        mean_ld_dur = str_to_float(r.get("mean_ld_duration_sec", ""))
        if math.isfinite(mean_ld_dur) and ld_i > 0:
            ld_dur_weight_sum += mean_ld_dur * ld_i
            ld_dur_weight += ld_i

        mean_dw_dur = str_to_float(r.get("mean_double_warning_duration_sec", ""))
        if math.isfinite(mean_dw_dur) and dw_i > 0:
            dw_dur_weight_sum += mean_dw_dur * dw_i
            dw_dur_weight += dw_i

        mean_int_gap = str_to_float(r.get("mean_interrupt_gap_sec", ""))
        if math.isfinite(mean_int_gap) and int_cnt_i > 0:
            int_gap_weight_sum += mean_int_gap * int_cnt_i
            int_gap_weight += int_cnt_i
        max_gap = str_to_float(r.get("max_interrupt_gap_sec", ""))
        if math.isfinite(max_gap):
            max_int_gap = max(max_int_gap, max_gap)

    denom_tp_fn = tp + fn
    denom_all = tp + fn + fp
    denom_prec = tp + fp
    tpr = (tp / denom_tp_fn) if denom_tp_fn > 0 else 0.0
    fnr = (fn / denom_tp_fn) if denom_tp_fn > 0 else 0.0
    fpr = (fp / denom_all) if denom_all > 0 else 0.0
    precision = (tp / denom_prec) if denom_prec > 0 else 0.0
    f1 = (2.0 * precision * tpr / (precision + tpr)) if (precision + tpr) > 0 else 0.0
    interruption_tp_ratio = (interrupted_tp / tp) if tp > 0 else 0.0

    def _wmean(sum_v: float, weight_v: float) -> str:
        if weight_v <= 0:
            return ""
        return fmt6(sum_v / weight_v)

    return {
        "generated_at_utc": first_generated_utc,
        "generated_at_cn": first_generated_cn,
        "kpi_tolerance_sec": tol_val,
        "radar_id": "0",
        "radar_direction": "ALL",
        "label_name": "ALL",
        "gt_intervals": str(gt),
        "pred_intervals": str(pred),
        "tp": str(tp),
        "fn": str(fn),
        "fp": str(fp),
        "tpr": fmt6(tpr),
        "fpr": fmt6(fpr),
        "fnr": fmt6(fnr),
        "precision": fmt6(precision),
        "f1": fmt6(f1),
        "mean_delay_sec": _wmean(delay_weight_sum, delay_weight),
        "min_delay_sec": (fmt6(min_delay) if math.isfinite(min_delay) else ""),
        "max_delay_sec": (fmt6(max_delay) if math.isfinite(max_delay) else ""),
        "mean_overlap_sec": _wmean(overlap_weight_sum, overlap_weight),
        "interrupted_tp": str(interrupted_tp),
        "interruption_count": str(interruption_count),
        "interruption_tp_ratio": fmt6(interruption_tp_ratio),
        "mean_interrupt_gap_sec": _wmean(int_gap_weight_sum, int_gap_weight),
        "max_interrupt_gap_sec": (fmt6(max_int_gap) if math.isfinite(max_int_gap) else ""),
        "mean_tp_duration_sec": _wmean(tp_dur_weight_sum, tp_dur_weight),
        "mean_fn_duration_sec": _wmean(fn_dur_weight_sum, fn_dur_weight),
        "mean_fp_duration_sec": _wmean(fp_dur_weight_sum, fp_dur_weight),
        "ea_count": str(ea_count),
        "mean_ea_duration_sec": _wmean(ea_dur_weight_sum, ea_dur_weight),
        "ed_count": str(ed_count),
        "mean_ed_duration_sec": _wmean(ed_dur_weight_sum, ed_dur_weight),
        "la_count": str(la_count),
        "mean_la_duration_sec": _wmean(la_dur_weight_sum, la_dur_weight),
        "ld_count": str(ld_count),
        "mean_ld_duration_sec": _wmean(ld_dur_weight_sum, ld_dur_weight),
        "double_warning_count": str(double_warning_count),
        "mean_double_warning_duration_sec": _wmean(dw_dur_weight_sum, dw_dur_weight),
        "ea_threshold_sec": ea_thr,
        "ed_threshold_sec": ed_thr,
        "la_threshold_sec": la_thr,
        "ld_threshold_sec": ld_thr,
    }


def export_batch_aggregate_report(output_dir: str, index_rows: List[Dict[str, str]]) -> Tuple[str, str]:
    ok_rows = [r for r in index_rows if trim(r.get("status", "")).upper() == "OK"]
    if not ok_rows:
        return "", ""

    bag_all_rows: List[Tuple[Dict[str, str], Dict[str, str]]] = []
    label_rows_all: List[Dict[str, str]] = []
    merged_summary_rows: List[Dict[str, str]] = []
    merged_event_rows: List[Dict[str, str]] = []

    for r in ok_rows:
        summary_rows = read_summary_rows_from_csv(r.get("summary_path", ""))
        if not summary_rows:
            summary_rows = read_sheet_rows_from_xlsx(r.get("xlsx_path", ""), "Summary")
        if not summary_rows:
            continue

        event_rows = read_summary_rows_from_csv(r.get("events_path", ""))
        if not event_rows:
            event_rows = read_sheet_rows_from_xlsx(r.get("xlsx_path", ""), "Events")

        bag_path = trim(r.get("bag_path", ""))
        if not bag_path and summary_rows:
            bag_path = trim(summary_rows[0].get("playback_bag_path", ""))
        bag_name = os.path.splitext(os.path.basename(bag_path))[0] if bag_path else ""
        if not bag_name and summary_rows:
            bag_name = os.path.splitext(os.path.basename(trim(summary_rows[0].get("gt_bag_path", ""))))[0]

        for sr in summary_rows:
            row = dict(sr)
            row["source_bag_name"] = bag_name
            row["source_bag_path"] = bag_path
            merged_summary_rows.append(row)
            if is_all_summary_row(sr):
                bag_all_rows.append(({"bag_path": bag_path}, row))
            else:
                label_rows_all.append(row)

        for er in event_rows:
            row = dict(er)
            row["source_bag_name"] = bag_name
            row["source_bag_path"] = bag_path
            merged_event_rows.append(row)

    if not bag_all_rows or (not merged_summary_rows):
        return "", ""

    def _ordered_headers(rows: List[Dict[str, str]], front: List[str]) -> List[str]:
        keys: List[str] = []
        seen = set()
        for k in front:
            if k not in seen:
                keys.append(k)
                seen.add(k)
        for row in rows:
            for k in row.keys():
                if k not in seen:
                    keys.append(k)
                    seen.add(k)
        return keys

    batch_summary_csv = os.path.join(output_dir, "batch_adas_kpi_summary.csv")
    batch_events_csv = os.path.join(output_dir, "batch_adas_kpi_summary_events.csv")
    summary_headers = _ordered_headers(merged_summary_rows, ["source_bag_name", "source_bag_path"])
    event_headers = _ordered_headers(merged_event_rows, ["source_bag_name", "source_bag_path"])

    with open(batch_summary_csv, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=summary_headers)
        w.writeheader()
        for row in merged_summary_rows:
            w.writerow({h: row.get(h, "") for h in summary_headers})

    with open(batch_events_csv, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=event_headers)
        w.writeheader()
        for row in merged_event_rows:
            w.writerow({h: row.get(h, "") for h in event_headers})

    batch_report_xlsx = os.path.join(output_dir, "batch_adas_kpi_report.xlsx")
    export_excel_report_from_csv(
        excel_path=batch_report_xlsx,
        summary_csv_path=batch_summary_csv,
        events_csv_path=batch_events_csv,
        report_note="Batch merged report: all bags combined into per-bag style sheets.",
    )

    batch_overall_row = build_batch_aggregate_row([x[1] for x in bag_all_rows])
    by_label: Dict[Tuple[int, str, str], List[Dict[str, str]]] = {}
    for r in label_rows_all:
        rid = str_to_int(r.get("radar_id", "0"))
        rdir = trim(r.get("radar_direction", ""))
        lname = trim(r.get("label_name", ""))
        by_label.setdefault((rid, rdir, lname), []).append(r)
    label_agg_rows: List[Dict[str, str]] = []
    for k in sorted(by_label.keys()):
        rr = build_batch_aggregate_row(by_label[k])
        rr["radar_id"] = str(k[0])
        rr["radar_direction"] = k[1]
        rr["label_name"] = k[2]
        label_agg_rows.append(rr)

    csv_path = os.path.join(output_dir, "batch_adas_kpi_overall.csv")
    csv_headers = [
        "scope",
        "bags",
        "bag_name",
        "bag_path",
        "radar_id",
        "radar_direction",
        "label_name",
        "gt_intervals",
        "pred_intervals",
        "tp",
        "fn",
        "fp",
        "tpr",
        "fpr",
        "fnr",
        "precision",
        "f1",
        "mean_delay_sec",
        "min_delay_sec",
        "max_delay_sec",
        "mean_overlap_sec",
        "interrupted_tp",
        "interruption_count",
        "interruption_tp_ratio",
        "mean_interrupt_gap_sec",
        "max_interrupt_gap_sec",
        "mean_tp_duration_sec",
        "mean_fn_duration_sec",
        "mean_fp_duration_sec",
        "ea_count",
        "mean_ea_duration_sec",
        "ed_count",
        "mean_ed_duration_sec",
        "la_count",
        "mean_la_duration_sec",
        "ld_count",
        "mean_ld_duration_sec",
        "double_warning_count",
        "mean_double_warning_duration_sec",
        "kpi_tolerance_sec",
        "ea_threshold_sec",
        "ed_threshold_sec",
        "la_threshold_sec",
        "ld_threshold_sec",
        "generated_at_cn",
    ]
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(csv_headers)
        for idx_row, all_row in bag_all_rows:
            bag_name = os.path.splitext(os.path.basename(idx_row.get("bag_path", "")))[0]
            row = {
                "scope": "BAG",
                "bags": "1",
                "bag_name": bag_name,
                "bag_path": idx_row.get("bag_path", ""),
            }
            row.update(all_row)
            w.writerow([row.get(h, "") for h in csv_headers])

        row = {"scope": "BATCH_ALL", "bags": str(len(bag_all_rows)), "bag_name": "ALL", "bag_path": ""}
        row.update(batch_overall_row)
        w.writerow([row.get(h, "") for h in csv_headers])
        for rr in label_agg_rows:
            row = {"scope": "BATCH_LABEL", "bags": str(len(bag_all_rows)), "bag_name": "ALL", "bag_path": ""}
            row.update(rr)
            w.writerow([row.get(h, "") for h in csv_headers])

    xlsx_path = batch_report_xlsx
    try:
        from openpyxl import load_workbook

        def _sheet_desc_cn(sheet_name: str) -> Tuple[str, str]:
            desc = {
                "ReportGuide_CN": ("报告阅读导航", "报告入口：先看子表用途，再看字段中文释义与示例。"),
                "Summary": ("汇总统计", "按来源bag+雷达+标签聚合后的核心KPI结果。"),
                "Events": ("事件明细", "逐事件明细（TP/FP/FN）及时序偏差/中断信息。"),
                "TP": ("真阳性明细", "从Events筛选 event_type=TP。"),
                "FP": ("误报明细", "从Events筛选 event_type=FP。"),
                "FN": ("漏报明细", "从Events筛选 event_type=FN。"),
                "EA": ("提前激活明细", "从Events筛选 is_ea=1。"),
                "ED": ("提前消失明细", "从Events筛选 is_ed=1。"),
                "LA": ("延迟激活明细", "从Events筛选 is_la=1。"),
                "LD": ("延迟消失明细", "从Events筛选 is_ld=1。"),
                "DoubleWarnings": ("双告警/中断明细", "从Events筛选 is_interrupted=1。"),
                "AdditionalStats_5G": ("五代附加统计映射", "对齐五代统计口径的汇总字段。"),
                "KPI_MetricDictionary": ("指标公式字典", "核心指标的公式/判定规则说明。"),
                "Template5G_Main": ("五代风格总览模板", "可交付/评审常用的汇总模板视图。"),
                "Template5G_Events": ("五代风格事件模板", "可交付/评审常用的事件模板视图。"),
                "Template5G_ManualGuide": ("人工填写指南", "人工结论字段定义与填写示例。"),
                "Meta": ("单报告元信息", "当前报告来源路径、生成时间、说明。"),
                "BatchOverall": ("整批总览", "对全部bag聚合后的总KPI。"),
                "BagOverview": ("单bag总览", "每个bag一行的ALL口径总览。"),
                "LabelAggregate": ("跨bag标签聚合", "按(雷达方向+标签)跨bag聚合。"),
                "BatchMeta": ("批次元信息", "批次输入索引与输出文件路径等信息。"),
            }
            return desc.get(sheet_name, ("子表说明", "见表头字段定义。"))

        def _field_cn_desc(field_name: str) -> Tuple[str, str, str]:
            mapping = {
                "source_bag_name": ("来源bag名", "该行数据来自哪个bag（去扩展名）。", "corner_radar_net_2026-04-01-13-58-48_0"),
                "source_bag_path": ("来源bag路径", "该行数据来源bag的绝对路径。", "/data/case_xxx.bag"),
                "generated_at_utc": ("UTC生成时间", "报告生成时间（UTC时区）。", "2026-04-24T09:42:02Z"),
                "generated_at_cn": ("中国时区生成时间", "报告生成时间（UTC+8）。", "2026-04-24T17:42:02+08:00"),
                "label_csv_path": ("标注CSV路径", "该行对应GT标注CSV文件路径。", "/data/case_xxx_corner_radar_gt.csv"),
                "playback_bag_path": ("回放bag路径", "用于回灌并统计KPI的bag路径。", "/data/case_xxx.bag"),
                "gt_bag_path": ("GT记录bag路径", "GT文件内部记录的bag路径字段。", "/data/case_xxx.bag"),
                "kpi_tolerance_sec": ("KPI容差秒", "当前strict-overlap口径固定写0。", "0.000000"),
                "radar_id": ("雷达ID", "角雷达编号：1/2/3/4；0通常表示ALL。", "3"),
                "radar_direction": ("雷达方向", "雷达方向：FL/FR/RL/RR/ALL。", "RL"),
                "label_name": ("报警标签", "报警类型标签名称。", "BSD_L"),
                "gt_intervals": ("GT区间数", "该聚合范围内GT区间条数。", "12"),
                "pred_intervals": ("预测区间数", "该聚合范围内预测区间条数。", "14"),
                "tp": ("TP数", "与GT有严格重叠(overlap>0)的事件数。", "9"),
                "fn": ("FN数", "GT中没有任何预测重叠的事件数。", "3"),
                "fp": ("FP数", "预测中不与任何GT重叠的事件数。", "5"),
                "tpr": ("召回率", "tp/(tp+fn)。", "0.750000"),
                "fpr": ("误报率", "fp/(tp+fn+fp)。", "0.294118"),
                "fnr": ("漏报率", "fn/(tp+fn)。", "0.250000"),
                "precision": ("精确率", "tp/(tp+fp)。", "0.642857"),
                "f1": ("F1分数", "2*precision*tpr/(precision+tpr)。", "0.692308"),
                "mean_delay_sec": ("平均延迟秒", "TP事件 delay_sec 的平均值。", "0.183333"),
                "min_delay_sec": ("最小延迟秒", "TP事件 delay_sec 的最小值。", "-0.050000"),
                "max_delay_sec": ("最大延迟秒", "TP事件 delay_sec 的最大值。", "0.420000"),
                "mean_overlap_sec": ("平均重叠秒", "TP事件 overlap_sec 的平均值。", "1.247500"),
                "interrupted_tp": ("中断TP数", "TP中 pred_segments>1 的事件数。", "2"),
                "interruption_count": ("中断次数", "中断总次数（分段数-1后聚合）。", "3"),
                "interruption_tp_ratio": ("中断TP占比", "interrupted_tp/tp。", "0.222222"),
                "mean_interrupt_gap_sec": ("平均中断空隙秒", "中断空隙列表展开后的平均值。", "0.310000"),
                "max_interrupt_gap_sec": ("最大中断空隙秒", "中断空隙列表展开后的最大值。", "0.800000"),
                "mean_tp_duration_sec": ("TP平均时长秒", "TP事件 pred_active_duration_sec 的平均值。", "2.000000"),
                "mean_fn_duration_sec": ("FN平均时长秒", "FN事件 gt_duration_sec 的平均值。", "1.500000"),
                "mean_fp_duration_sec": ("FP平均时长秒", "FP事件 pred_duration_sec 的平均值。", "1.000000"),
                "ea_count": ("EA数量", "is_ea==1 的事件数。", "1"),
                "mean_ea_duration_sec": ("EA平均时长秒", "EA事件 ea_duration_sec 平均值。", "0.400000"),
                "ed_count": ("ED数量", "is_ed==1 的事件数。", "2"),
                "mean_ed_duration_sec": ("ED平均时长秒", "ED事件 ed_duration_sec 平均值。", "0.400000"),
                "la_count": ("LA数量", "is_la==1 的事件数。", "1"),
                "mean_la_duration_sec": ("LA平均时长秒", "LA事件 la_duration_sec 平均值。", "0.600000"),
                "ld_count": ("LD数量", "is_ld==1 的事件数。", "0"),
                "mean_ld_duration_sec": ("LD平均时长秒", "LD事件 ld_duration_sec 平均值。分母为0时空。", ""),
                "double_warning_count": ("双告警次数", "中断TP事件数（当前口径等价 interrupted_tp）。", "1"),
                "mean_double_warning_duration_sec": ("双告警平均时长秒", "中断TP的 double_warning_duration_sec 平均值。", "0.700000"),
                "ea_threshold_sec": ("EA阈值秒", "EA判定阈值。", "0.250000"),
                "ed_threshold_sec": ("ED阈值秒", "ED判定阈值。", "0.250000"),
                "la_threshold_sec": ("LA阈值秒", "LA判定阈值。", "0.250000"),
                "ld_threshold_sec": ("LD阈值秒", "LD判定阈值。", "0.250000"),
                "event_type": ("事件类型", "事件类别：TP/FP/FN。", "TP"),
                "gt_start_abs": ("GT开始绝对秒", "GT区间起点ROS绝对秒。", "1711941368.123456"),
                "gt_start_cn": ("GT开始北京时间", "由绝对秒换算的人可读北京时间。", "2026-04-01 13:56:08.123"),
                "gt_end_abs": ("GT结束绝对秒", "GT区间终点ROS绝对秒。", "1711941370.123456"),
                "gt_end_cn": ("GT结束北京时间", "由绝对秒换算的人可读北京时间。", "2026-04-01 13:56:10.123"),
                "pred_start_abs": ("预测主段开始绝对秒", "TP/FP主段开始绝对秒。", "1711941368.300000"),
                "pred_start_cn": ("预测主段开始北京时间", "由绝对秒换算的人可读北京时间。", "2026-04-01 13:56:08.300"),
                "pred_end_abs": ("预测主段结束绝对秒", "TP/FP主段结束绝对秒。", "1711941369.000000"),
                "pred_end_cn": ("预测主段结束北京时间", "由绝对秒换算的人可读北京时间。", "2026-04-01 13:56:09.000"),
                "pred_first_start_abs": ("预测首段开始绝对秒", "多段预测中第一段起点。", "1711941368.300000"),
                "pred_first_start_cn": ("预测首段开始北京时间", "首段起点的人可读北京时间。", "2026-04-01 13:56:08.300"),
                "pred_last_end_abs": ("预测末段结束绝对秒", "多段预测中最后一段终点。", "1711941369.800000"),
                "pred_last_end_cn": ("预测末段结束北京时间", "末段终点的人可读北京时间。", "2026-04-01 13:56:09.800"),
                "gt_start_rel": ("GT开始相对秒", "相对bag起始时刻的秒数。", "20.642098"),
                "gt_end_rel": ("GT结束相对秒", "相对bag起始时刻的秒数。", "22.642098"),
                "pred_start_rel": ("预测主段开始相对秒", "相对bag起始时刻的秒数。", "20.820000"),
                "pred_end_rel": ("预测主段结束相对秒", "相对bag起始时刻的秒数。", "21.520000"),
                "pred_first_start_rel": ("预测首段开始相对秒", "相对bag起始时刻的秒数。", "20.820000"),
                "pred_last_end_rel": ("预测末段结束相对秒", "相对bag起始时刻的秒数。", "22.100000"),
                "delay_sec": ("延迟秒", "pred_start_abs-gt_start_abs（仅TP有意义）。", "0.177902"),
                "overlap_sec": ("重叠秒", "GT与预测重叠时长（仅TP有意义）。", "1.200000"),
                "gt_duration_sec": ("GT时长秒", "gt_end_abs-gt_start_abs。", "2.000000"),
                "pred_duration_sec": ("预测主段时长秒", "pred_end_abs-pred_start_abs。", "0.700000"),
                "pred_active_duration_sec": ("预测有效激活总时长秒", "多段预测各段时长求和（不含中断空隙）。", "1.300000"),
                "pred_span_duration_sec": ("预测首末跨度时长秒", "pred_last_end_abs-pred_first_start_abs（含空隙）。", "1.800000"),
                "pred_segments": ("预测分段数", "GT重叠后合并得到的预测段数量。", "2"),
                "interruption_gaps_sec": ("中断空隙列表秒", "相邻预测段之间空隙；分号分隔。", "0.500000;0.300000"),
                "is_interrupted": ("是否中断", "1表示该TP存在中断（pred_segments>1）。", "1"),
                "start_offset_sec": ("起点偏差秒", "pred_first_start_abs-gt_start_abs。", "0.200000"),
                "end_offset_sec": ("终点偏差秒", "pred_last_end_abs-gt_end_abs。", "-0.100000"),
                "ea_duration_sec": ("提前激活时长秒", "max(0,-start_offset_sec)。", "0.000000"),
                "ed_duration_sec": ("提前消失时长秒", "max(0,-end_offset_sec)。", "0.100000"),
                "la_duration_sec": ("延迟激活时长秒", "max(0,start_offset_sec)。", "0.200000"),
                "ld_duration_sec": ("延迟消失时长秒", "max(0,end_offset_sec)。", "0.000000"),
                "is_ea": ("是否EA", "ea_duration_sec>ea_threshold_sec。", "0"),
                "is_ed": ("是否ED", "ed_duration_sec>ed_threshold_sec。", "0"),
                "is_la": ("是否LA", "la_duration_sec>la_threshold_sec。", "0"),
                "is_ld": ("是否LD", "ld_duration_sec>ld_threshold_sec。", "0"),
                "double_warning_duration_sec": ("双告警持续时长秒", "pred_span_duration_sec-pred_active_duration_sec。", "0.500000"),
                "mean_duration_tp_sec": ("TP平均时长秒(五代映射)", "与 Summary.mean_tp_duration_sec 同口径。", "2.000000"),
                "mean_duration_fn_sec": ("FN平均时长秒(五代映射)", "与 Summary.mean_fn_duration_sec 同口径。", "1.500000"),
                "mean_duration_fp_sec": ("FP平均时长秒(五代映射)", "与 Summary.mean_fp_duration_sec 同口径。", "1.000000"),
                "mean_duration_ea_sec": ("EA平均时长秒(五代映射)", "与 Summary.mean_ea_duration_sec 同口径。", "0.400000"),
                "mean_duration_ed_sec": ("ED平均时长秒(五代映射)", "与 Summary.mean_ed_duration_sec 同口径。", "0.400000"),
                "mean_duration_la_sec": ("LA平均时长秒(五代映射)", "与 Summary.mean_la_duration_sec 同口径。", "0.600000"),
                "mean_duration_ld_sec": ("LD平均时长秒(五代映射)", "与 Summary.mean_ld_duration_sec 同口径。分母0时空。", ""),
                "mean_duration_double_warning_sec": ("双告警平均时长秒(五代映射)", "与 Summary.mean_double_warning_duration_sec 同口径。", "0.700000"),
                "MetricKey(指标键)": ("指标键", "指标英文键名，用于程序和字段映射。", "tp"),
                "ChineseName(中文名称)": ("中文名称", "指标中文名，用于评审/客户阅读。", "真阳性数"),
                "Level(层级)": ("层级", "Event/Summary 等指标生效层级。", "Summary"),
                "SheetSource(来源子表)": ("来源子表", "该指标原始来源子表。", "Summary"),
                "FormulaOrRule(计算公式/判定规则)": ("公式/规则", "指标计算公式或判定条件。", "tp/(tp+fn)"),
                "Notes(备注)": ("备注", "口径说明或边界条件。", "仅中断TP统计"),
                "SourceBagName(来源bag名)": ("来源bag名(模板)", "该模板行对应的bag名。", "corner_radar_net_..."),
                "SourceBagPath(来源bag路径)": ("来源bag路径(模板)", "该模板行对应bag绝对路径。", "/data/case_xxx.bag"),
                "ReactionType(反应类型)": ("反应类型", "通常对应报警标签类型。", "BSD_L"),
                "RadarDirection(雷达方向)": ("雷达方向", "FL/FR/RL/RR。", "RR"),
                "GT_Count(GT区间数)": ("GT区间数", "该行聚合范围内GT区间数量。", "12"),
                "Pred_Count(预测区间数)": ("预测区间数", "该行聚合范围内预测区间数量。", "14"),
                "Recall_TPR(召回率)": ("召回率", "tp/(tp+fn)。", "0.875"),
                "MeanTPDurSec(TP平均时长秒)": ("TP平均时长秒", "TP事件有效激活时长平均值。", "2.000000"),
                "MeanFNDurSec(FN平均时长秒)": ("FN平均时长秒", "FN事件GT时长平均值。", "1.500000"),
                "MeanFPDurSec(FP平均时长秒)": ("FP平均时长秒", "FP事件预测时长平均值。", "1.000000"),
                "EACount(提前激活数量)": ("EA数量", "is_ea==1 的事件数量。", "1"),
                "EDCount(提前消失数量)": ("ED数量", "is_ed==1 的事件数量。", "2"),
                "LACount(延迟激活数量)": ("LA数量", "is_la==1 的事件数量。", "1"),
                "LDCount(延迟消失数量)": ("LD数量", "is_ld==1 的事件数量。", "0"),
                "DoubleWarnCount(双告警数量)": ("双告警数量", "中断TP事件数量。", "1"),
                "ManualConclusion(人工结论)": ("人工结论", "人工最终结论（通过/关注/不通过等）。", "关注"),
                "ManualComment(人工备注)": ("人工备注", "对异常原因或场景补充说明。", "雨天反光导致误报"),
                "PER(算法评审)": ("算法评审", "算法侧评审结论。", "参数需优化"),
                "SIT(系统评审)": ("系统评审", "系统集成侧评审结论。", "时序正常"),
                "FCT(功能评审)": ("功能评审", "功能测试侧评审结论。", "场景覆盖完整"),
                "Category(分类)": ("分类", "问题分类标签。", "阈值/时序/标注"),
                "RecallPct(召回率百分比)": ("召回率百分比", "Excel公式列：Recall_TPR*100。", "87.5"),
                "EventType(事件类型)": ("事件类型(模板)", "TP/FP/FN。", "TP"),
                "GT_Start_Abs(真值开始绝对秒)": ("GT开始绝对秒", "真值区间开始绝对秒。", "1711941368.123456"),
                "GT_Start_CN(真值开始北京时间)": ("GT开始北京时间", "绝对秒转北京时间。", "2026-04-01 13:56:08.123"),
                "GT_End_Abs(真值结束绝对秒)": ("GT结束绝对秒", "真值区间结束绝对秒。", "1711941370.123456"),
                "GT_End_CN(真值结束北京时间)": ("GT结束北京时间", "绝对秒转北京时间。", "2026-04-01 13:56:10.123"),
                "Pred_Start_Abs(预测开始绝对秒)": ("预测开始绝对秒", "预测主段开始绝对秒。", "1711941368.300000"),
                "Pred_Start_CN(预测开始北京时间)": ("预测开始北京时间", "绝对秒转北京时间。", "2026-04-01 13:56:08.300"),
                "Pred_End_Abs(预测结束绝对秒)": ("预测结束绝对秒", "预测主段结束绝对秒。", "1711941369.000000"),
                "Pred_End_CN(预测结束北京时间)": ("预测结束北京时间", "绝对秒转北京时间。", "2026-04-01 13:56:09.000"),
                "DelaySec(延迟秒)": ("延迟秒", "pred_start_abs-gt_start_abs。", "0.177902"),
                "OverlapSec(重叠秒)": ("重叠秒", "GT与Pred重叠时长。", "1.200000"),
                "GT_DurationSec(真值时长秒)": ("GT时长秒", "gt_end_abs-gt_start_abs。", "2.000000"),
                "Pred_Active_DurationSec(预测有效时长秒)": ("预测有效时长秒", "多段预测各段时长之和。", "1.300000"),
                "Pred_Segments(预测分段数)": ("预测分段数", "重叠后预测分段数量。", "2"),
                "IsInterrupted(是否中断)": ("是否中断", "pred_segments>1 判为1。", "1"),
                "IsEA(是否提前激活)": ("是否EA", "ea_duration>ea_threshold 判为1。", "0"),
                "IsED(是否提前消失)": ("是否ED", "ed_duration>ed_threshold 判为1。", "0"),
                "IsLA(是否延迟激活)": ("是否LA", "la_duration>la_threshold 判为1。", "0"),
                "IsLD(是否延迟消失)": ("是否LD", "ld_duration>ld_threshold 判为1。", "0"),
                "ManualLabeling(人工标注结论)": ("人工标注结论", "人工复核结论字段。", "可接受"),
                "NotConsidered(是否不纳统)": ("是否不纳统", "该事件是否不计入KPI统计（0/1）。", "0"),
                "NotConsideredReason(不纳统原因)": ("不纳统原因", "若不纳统，填写理由。", "施工场景不在设计域"),
                "Field(字段)": ("字段名", "人工指南中的字段名。", "ManualConclusion"),
                "Type(类型)": ("类型", "字段类型（自动/人工填写）。", "Manual"),
                "Meaning(含义)": ("含义", "字段业务含义说明。", "人工最终结论"),
                "Example(示例)": ("示例", "字段填写示例。", "关注"),
                "scope": ("聚合层级", "BAG/BATCH_ALL/BATCH_LABEL。", "BATCH_LABEL"),
                "bags": ("bag数量", "该聚合行覆盖的bag数量。", "12"),
                "bag_name": ("bag名", "bag文件名（无扩展名）。", "corner_radar_net_..."),
                "bag_path": ("bag路径", "bag绝对路径（BagOverview常用）。", "/data/case_xxx.bag"),
                "key": ("键", "Meta子表键名。", "generated_at_cn"),
                "value": ("值", "Meta子表键值。", "2026-04-30T15:32:11+08:00"),
            }

            f = trim(field_name)
            if f in mapping:
                return mapping[f]

            if "(" in f and ")" in f:
                cn = trim(f.split("(", 1)[1].rsplit(")", 1)[0]) or f
                return (cn, "模板字段，详见Template5G与人工填写说明。", "")

            return (f, "字段含义待补充（可参考 Summary/Events 同名字段口径）。", "")

        wb = load_workbook(batch_report_xlsx)
        for sheet_name in ("BatchOverall", "BagOverview", "LabelAggregate", "BatchMeta", "ReportGuide_CN"):
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

        ws_overall = wb.create_sheet("BatchOverall")
        ws_overall.append(csv_headers)
        row = {"scope": "BATCH_ALL", "bags": str(len(bag_all_rows)), "bag_name": "ALL", "bag_path": ""}
        row.update(batch_overall_row)
        ws_overall.append([row.get(h, "") for h in csv_headers])

        ws_bag = wb.create_sheet("BagOverview")
        ws_bag.append(csv_headers)
        for idx_row, all_row in bag_all_rows:
            bag_name = os.path.splitext(os.path.basename(idx_row.get("bag_path", "")))[0]
            row = {
                "scope": "BAG",
                "bags": "1",
                "bag_name": bag_name,
                "bag_path": idx_row.get("bag_path", ""),
            }
            row.update(all_row)
            ws_bag.append([row.get(h, "") for h in csv_headers])

        ws_label = wb.create_sheet("LabelAggregate")
        ws_label.append(csv_headers)
        for rr in label_agg_rows:
            row = {"scope": "BATCH_LABEL", "bags": str(len(bag_all_rows)), "bag_name": "ALL", "bag_path": ""}
            row.update(rr)
            ws_label.append([row.get(h, "") for h in csv_headers])

        ws_meta = wb.create_sheet("BatchMeta")
        ws_meta.append(["key", "value"])
        ws_meta.append(["bags_ok", str(len(bag_all_rows))])
        ws_meta.append(["index_rows", str(len(index_rows))])
        ws_meta.append(["index_path", os.path.join(output_dir, "batch_kpi_index.csv")])
        ws_meta.append(["batch_csv_path", csv_path])
        ws_meta.append(["batch_summary_csv_path", batch_summary_csv])
        ws_meta.append(["batch_events_csv_path", batch_events_csv])
        ws_meta.append(["generated_at_cn", dt.datetime.now().astimezone(dt.timezone(dt.timedelta(hours=8))).isoformat()])

        ws_guide = wb.create_sheet("ReportGuide_CN")
        ws_guide.append(["Section(章节)", "Sheet(子表)", "Field(字段)", "ChineseName(中文名称)", "Meaning(含义)", "Example(示例)"])

        read_order = [
            "ReportGuide_CN",
            "BatchMeta",
            "BatchOverall",
            "BagOverview",
            "LabelAggregate",
            "Summary",
            "Events",
            "TP",
            "FP",
            "FN",
            "EA",
            "ED",
            "LA",
            "LD",
            "DoubleWarnings",
            "AdditionalStats_5G",
            "KPI_MetricDictionary",
            "Template5G_Main",
            "Template5G_Events",
            "Template5G_ManualGuide",
            "Meta",
        ]

        ws_guide.append(["SheetGuide(子表导览)", "", "", "", "推荐阅读顺序见下方行序。", ""])
        for s in read_order:
            if s not in wb.sheetnames and s != "ReportGuide_CN":
                continue
            cn, meaning = _sheet_desc_cn(s)
            ws_guide.append(["SheetGuide(子表导览)", s, "", cn, meaning, ""])

        ws_guide.append(["", "", "", "", "", ""])
        ws_guide.append(["FieldGuide(字段字典)", "", "", "", "按子表逐字段列出中英和含义。", ""])

        dict_sheets = [s for s in read_order if s in wb.sheetnames and s != "ReportGuide_CN"]
        seen = set()
        for s in dict_sheets:
            ws = wb[s]
            if ws.max_row < 1:
                continue
            headers = [trim(str(c.value)) if c.value is not None else "" for c in ws[1]]
            headers = [h for h in headers if h]
            if not headers:
                continue
            for h in headers:
                key = (s, h)
                if key in seen:
                    continue
                seen.add(key)
                cn, meaning, example = _field_cn_desc(h)
                ws_guide.append(["FieldGuide(字段字典)", s, h, cn, meaning, example])

        ws_guide.append(["", "", "", "", "", ""])
        ws_guide.append(["KeyGuide(常见键说明)", "Meta/BatchMeta", "key", "键名", "常见键及含义如下。", ""])
        key_guides = [
            ("summary_csv_path", "Summary CSV路径", "/output/case_xxx_adas_kpi_summary.csv"),
            ("events_csv_path", "Events CSV路径", "/output/case_xxx_adas_kpi_summary_events.csv"),
            ("generated_at_utc", "报告UTC生成时间", "2026-04-24T09:42:02Z"),
            ("generated_at_cn", "报告中国时区生成时间", "2026-04-24T17:42:02+08:00"),
            ("abs_time_human_timezone", "人可读时间时区", "Asia/Shanghai (UTC+08:00)"),
            ("abs_time_human_format", "人可读时间格式", "YYYY-MM-DD HH:MM:SS.mmm"),
            ("note", "补充说明", "Batch merged report ..."),
            ("bags_ok", "成功bag数量", "12"),
            ("index_rows", "索引总行数", "12"),
            ("index_path", "批次索引文件路径", "/output/batch_kpi_index.csv"),
            ("batch_csv_path", "批次总览CSV路径", "/output/batch_adas_kpi_overall.csv"),
            ("batch_summary_csv_path", "批次Summary合并CSV路径", "/output/batch_adas_kpi_summary.csv"),
            ("batch_events_csv_path", "批次Events合并CSV路径", "/output/batch_adas_kpi_summary_events.csv"),
        ]
        for k, meaning, ex in key_guides:
            ws_guide.append(["KeyGuide(常见键说明)", "Meta/BatchMeta", k, k, meaning, ex])

        # Place guide sheet at the front to fit common customer-reading habits.
        guide_idx = wb.index(ws_guide)
        if guide_idx > 0:
            wb.move_sheet(ws_guide, offset=-guide_idx)

        wb.save(xlsx_path)
    except Exception:
        xlsx_path = batch_report_xlsx if os.path.isfile(batch_report_xlsx) else ""

    return csv_path, xlsx_path


def export_batch_aggregate_report_from_output_dir(output_dir: str) -> Tuple[str, str]:
    if not output_dir or (not os.path.isdir(output_dir)):
        return "", ""
    rows: List[Dict[str, str]] = []
    seen_bases = set()
    for fn in sorted(os.listdir(output_dir)):
        if not fn.endswith("_adas_kpi_summary.csv"):
            continue
        if fn.startswith("batch_"):
            continue
        summary_path = os.path.join(output_dir, fn)
        base = fn[:-len("_adas_kpi_summary.csv")]
        seen_bases.add(base)
        xlsx_path = os.path.join(output_dir, f"{base}_adas_kpi_report.xlsx")
        rows.append(
            {
                "status": "OK",
                "summary_path": summary_path,
                "events_path": os.path.join(output_dir, f"{base}_adas_kpi_summary_events.csv"),
                "xlsx_path": (xlsx_path if os.path.isfile(xlsx_path) else ""),
                "bag_path": "",
                "csv_path": "",
                "detail": "",
            }
        )
    for fn in sorted(os.listdir(output_dir)):
        if not fn.endswith("_adas_kpi_report.xlsx"):
            continue
        if fn.startswith("batch_"):
            continue
        base = fn[:-len("_adas_kpi_report.xlsx")]
        if base in seen_bases:
            continue
        xlsx_path = os.path.join(output_dir, fn)
        rows.append(
            {
                "status": "OK",
                "summary_path": os.path.join(output_dir, f"{base}_adas_kpi_summary.csv"),
                "events_path": os.path.join(output_dir, f"{base}_adas_kpi_summary_events.csv"),
                "xlsx_path": xlsx_path,
                "bag_path": "",
                "csv_path": "",
                "detail": "",
            }
        )
    if not rows:
        return "", ""
    return export_batch_aggregate_report(output_dir, rows)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Corner radar KPI batch tool (bag + GT csv)")

    mode = p.add_mutually_exclusive_group(required=True)
    mode.add_argument("--input-dir", help="Directory mode: recursively match bag+csv pairs")
    mode.add_argument("--bag", help="Single mode: input bag path")

    p.add_argument("--csv", help="Single mode: GT CSV path (required with --bag)")
    p.add_argument("--output-dir", default="", help="Output directory. Default: <base>/kpi_reports_YYYYmmdd_HHMMSS")
    p.add_argument("--warning-topic", default="/corner_radar/warning_status_raw", help="Warning topic in bag")
    p.add_argument("--lgu-prefix", default="/wf/corner_radar/lgu_data_", help="LGU topic prefix")
    p.add_argument("--tol", type=float, default=0.0, help="Deprecated. Ignored in strict-overlap PM mode.")
    p.add_argument(
        "--report-format",
        choices=["csv", "xlsx", "both"],
        default="both",
        help="Output format: csv, xlsx, or both. Default: both",
    )
    p.add_argument("--ea-threshold-sec", type=float, default=K_DEFAULT_EA_THRESHOLD_SEC, help="EA threshold in seconds")
    p.add_argument("--ed-threshold-sec", type=float, default=K_DEFAULT_ED_THRESHOLD_SEC, help="ED threshold in seconds")
    p.add_argument("--la-threshold-sec", type=float, default=K_DEFAULT_LA_THRESHOLD_SEC, help="LA threshold in seconds")
    p.add_argument("--ld-threshold-sec", type=float, default=K_DEFAULT_LD_THRESHOLD_SEC, help="LD threshold in seconds")
    p.add_argument("--verbose", action="store_true", help="Print extra stats")

    args = p.parse_args()
    if args.bag and not args.csv:
        p.error("--csv is required when --bag is used")
    return args


def main() -> int:
    args = parse_args()

    rows: List[Dict[str, str]] = []
    success = 0
    failed = 0

    if args.input_dir:
        input_dir = os.path.abspath(args.input_dir)
        jobs, missing = collect_jobs(input_dir)
        if not jobs:
            print("[ERROR] No matched bag+csv jobs found.", file=sys.stderr)
            if missing:
                print("[INFO] Missing examples:", file=sys.stderr)
                for m in missing[:10]:
                    print("  -", m, file=sys.stderr)
            return 2

        out_dir = ensure_output_dir(args.output_dir if args.output_dir else "", input_dir)
        print(f"[INFO] Batch jobs: {len(jobs)}")
        print(f"[INFO] Output dir: {out_dir}")
        if missing:
            print(f"[WARN] Missing csv for {len(missing)} bag(s). Skipped.")

        for i, job in enumerate(jobs, start=1):
            print(f"[INFO] ({i}/{len(jobs)}) bag={job.bag_path}")
            row = {
                "bag_path": job.bag_path,
                "csv_path": job.csv_path,
                "status": "FAILED",
                "summary_path": "",
                "events_path": "",
                "xlsx_path": "",
                "detail": "",
            }
            try:
                ok, detail, sum_path, ev_path, xlsx_path, _warns = run_one_job(
                    bag_path=job.bag_path,
                    csv_path=job.csv_path,
                    output_dir=out_dir,
                    warning_topic=args.warning_topic,
                    lgu_prefix=args.lgu_prefix,
                    tol_sec=float(args.tol),
                    ea_threshold_sec=float(args.ea_threshold_sec),
                    ed_threshold_sec=float(args.ed_threshold_sec),
                    la_threshold_sec=float(args.la_threshold_sec),
                    ld_threshold_sec=float(args.ld_threshold_sec),
                    report_format=args.report_format,
                    verbose=bool(args.verbose),
                )
                row["status"] = "OK" if ok else "FAILED"
                row["summary_path"] = sum_path
                row["events_path"] = ev_path
                row["xlsx_path"] = xlsx_path
                row["detail"] = detail
                if ok:
                    success += 1
                else:
                    failed += 1
            except Exception as exc:
                failed += 1
                row["detail"] = f"FAILED: {exc}"
            rows.append(row)

        index_path = os.path.join(out_dir, "batch_kpi_index.csv")
        write_index(index_path, rows)
        batch_csv_path, batch_xlsx_path = export_batch_aggregate_report(out_dir, rows)
        print(f"[INFO] Done. success={success}, failed={failed}")
        print(f"[INFO] Index: {index_path}")
        if batch_csv_path:
            print(f"[INFO] Batch CSV : {batch_csv_path}")
        if batch_xlsx_path:
            print(f"[INFO] Batch XLSX: {batch_xlsx_path}")
        return 0 if failed == 0 else 1

    # Single mode
    bag_path = os.path.abspath(args.bag)
    csv_path = os.path.abspath(args.csv)
    base_dir = os.path.dirname(bag_path)
    out_dir = ensure_output_dir(args.output_dir if args.output_dir else "", base_dir)

    row = {
        "bag_path": bag_path,
        "csv_path": csv_path,
        "status": "FAILED",
        "summary_path": "",
        "events_path": "",
        "xlsx_path": "",
        "detail": "",
    }

    try:
        ok, detail, sum_path, ev_path, xlsx_path, _warns = run_one_job(
            bag_path=bag_path,
            csv_path=csv_path,
            output_dir=out_dir,
            warning_topic=args.warning_topic,
            lgu_prefix=args.lgu_prefix,
            tol_sec=float(args.tol),
            ea_threshold_sec=float(args.ea_threshold_sec),
            ed_threshold_sec=float(args.ed_threshold_sec),
            la_threshold_sec=float(args.la_threshold_sec),
            ld_threshold_sec=float(args.ld_threshold_sec),
            report_format=args.report_format,
            verbose=bool(args.verbose),
        )
        row["status"] = "OK" if ok else "FAILED"
        row["summary_path"] = sum_path
        row["events_path"] = ev_path
        row["xlsx_path"] = xlsx_path
        row["detail"] = detail
        success = 1 if ok else 0
        failed = 0 if ok else 1
    except Exception as exc:
        failed = 1
        success = 0
        row["detail"] = f"FAILED: {exc}"

    rows.append(row)
    index_path = os.path.join(out_dir, "batch_kpi_index.csv")
    write_index(index_path, rows)
    batch_csv_path, batch_xlsx_path = export_batch_aggregate_report_from_output_dir(out_dir)

    print(f"[INFO] Done. success={success}, failed={failed}")
    print(f"[INFO] Summary: {row['summary_path']}")
    print(f"[INFO] Events : {row['events_path']}")
    print(f"[INFO] XLSX   : {row.get('xlsx_path', '')}")
    print(f"[INFO] Index  : {index_path}")
    if batch_csv_path:
        print(f"[INFO] Batch CSV : {batch_csv_path}")
    if batch_xlsx_path:
        print(f"[INFO] Batch XLSX: {batch_xlsx_path}")

    return 0 if failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
